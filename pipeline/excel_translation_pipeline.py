# pipeline/excel_translation_pipeline.py  By AI-Transtools
# Unified Excel translation pipeline supporting openpyxl, xlwings, and bilingual modes
import os
import json
import re
import tempfile
import shutil
from datetime import datetime
from zipfile import ZipFile
from lxml import etree
from typing import Dict, List, Any, Tuple, Set
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries
from .skip_pipeline import should_translate
from config.log_config import app_logger


def extract_excel_content_to_json(file_path, temp_dir, use_xlwings=False):
    """
    Extract content from Excel file to JSON.

    Args:
        file_path: Path to the Excel file
        temp_dir: Temporary directory for processing
        use_xlwings: If True, use xlwings for extraction (requires Excel installed)
    """
    if use_xlwings:
        return _extract_with_xlwings(file_path, temp_dir)
    else:
        return _extract_with_openpyxl(file_path, temp_dir)


def write_translated_content_to_excel(file_path, original_json_path, translated_json_path,
                                       result_dir, use_xlwings=False, bilingual_mode=False, src_lang=None, dst_lang=None):
    """
    Write translated content back to Excel file.

    Args:
        file_path: Path to the original Excel file
        original_json_path: Path to the original JSON data
        translated_json_path: Path to the translated JSON data
        result_dir: Directory to save the result
        use_xlwings: If True, use xlwings for writing (requires Excel installed)
        bilingual_mode: If True, format content as bilingual (original + translated)
        src_lang: Source language code (e.g., 'zh')
        dst_lang: Target language code (e.g., 'ja')
    """
    if use_xlwings:
        return _write_with_xlwings(file_path, original_json_path, translated_json_path,
                                   result_dir, bilingual_mode, src_lang, dst_lang)
    else:
        return _write_with_openpyxl(file_path, original_json_path, translated_json_path, result_dir, bilingual_mode, src_lang, dst_lang)


# ============================================================================
# SHARED UTILITIES
# ============================================================================

def sanitize_sheet_name(sheet_name):
    """
    Clean sheet name by removing/replacing invalid characters.
    Excel doesn't allow these characters in sheet names: / \\ ? * [ ] :
    """
    invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
    sanitized_name = sheet_name
    for char in invalid_chars:
        sanitized_name = sanitized_name.replace(char, '-')

    # Excel sheet names are limited to 31 characters
    if len(sanitized_name) > 31:
        sanitized_name = sanitized_name[:31]

    # Ensure the sheet name is not empty
    if not sanitized_name.strip():
        sanitized_name = "Sheet"

    return sanitized_name


def _format_bilingual_text(original_text: str, translated_text: str, content_type: str = "cell") -> str:
    """
    Format original and translated text for bilingual display.

    Args:
        original_text: The original text
        translated_text: The translated text
        content_type: Type of content ('cell', 'textbox', 'smartart', 'sheet_name', 'drawing')

    Returns:
        Formatted bilingual text
    """
    # Clean up text by restoring line breaks
    original_clean = original_text.replace("␊", "\n").replace("␍", "\r")
    translated_clean = translated_text.replace("␊", "\n").replace("␍", "\r")

    # Format based on content type
    if content_type == "sheet_name":
        # For sheet names, use "Original (Translation)" format
        if original_clean.strip() == translated_clean.strip():
            return original_clean  # No need for duplication if they're the same
        return f"{original_clean} ({translated_clean})"

    elif content_type in ["cell", "textbox", "smartart", "drawing"]:
        # For cells, textboxes, SmartArt, and drawings, use line-separated format
        if original_clean.strip() == translated_clean.strip():
            return original_clean  # No need for duplication if they're the same

        # Use simple line-separated format for all content
        return f"{original_clean}\n{translated_clean}"

    else:
        # Default format
        return f"{original_clean}\n{translated_clean}"


# ============================================================================
# OPENPYXL MODE - EXTRACTION
# ============================================================================

def _extract_with_openpyxl(file_path, temp_dir):
    """Extract Excel content using openpyxl library."""
    workbook = load_workbook(file_path)
    cell_data = []
    count = 0

    # Add sheet names to the extraction process
    for sheet_name in workbook.sheetnames:
        # Add sheet name as a special entry if it should be translated
        if should_translate(sheet_name):
            count += 1
            sheet_info = {
                "count_src": count,
                "sheet": "SHEET_NAME",  # Special marker to identify sheet names
                "row": 0,               # Use 0 to indicate it's a sheet name, not a cell
                "column": 0,            # Use 0 to indicate it's a sheet name, not a cell
                "value": sheet_name,
                "is_merged": False,
                "is_sheet_name": True   # Flag to identify this as a sheet name entry
            }
            cell_data.append(sheet_info)

        sheet = workbook[sheet_name]
        merged_cells_ranges = sheet.merged_cells.ranges

        for row in sheet.iter_rows():
            for cell in row:
                # Skip cells with no value
                if cell.value is None:
                    continue

                # Skip datetime objects
                if isinstance(cell.value, datetime):
                    continue

                # Skip formulas (cells that start with '=')
                if isinstance(cell.value, str) and cell.value.strip().startswith('='):
                    continue

                # Skip cells that shouldn't be translated
                if not should_translate(str(cell.value)):
                    continue

                # Skip merged cells (except the top-left cell of the merge)
                if isinstance(cell, MergedCell):
                    continue

                is_merged_cell = False
                for merged_range in merged_cells_ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    if cell.row == min_row and cell.column == min_col:
                        is_merged_cell = True
                        break

                # Convert datetime values to string
                cell_value = str(cell.value).replace("\n", "␊").replace("\r", "␍")
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.isoformat()

                count += 1
                cell_info = {
                    "count_src": count,
                    "sheet": sheet_name,
                    "row": cell.row,
                    "column": cell.column,
                    "value": cell_value,
                    "is_merged": is_merged_cell,
                    "is_sheet_name": False  # Regular cell, not a sheet name
                }
                cell_data.append(cell_info)

    filename = os.path.splitext(os.path.basename(file_path))[0]
    temp_folder = os.path.join(temp_dir, filename)
    os.makedirs(temp_folder, exist_ok=True)
    json_path = os.path.join(temp_folder, "src.json")

    with open(json_path, "w", encoding="utf-8") as json_file:
        json.dump(cell_data, json_file, ensure_ascii=False, indent=4)

    return json_path


# ============================================================================
# OPENPYXL MODE - WRITING
# ============================================================================

def _write_with_openpyxl(file_path, original_json_path, translated_json_path, result_dir, bilingual_mode=False, src_lang=None, dst_lang=None):
    """Write translated content using openpyxl library."""
    workbook = load_workbook(file_path)

    # Load original JSON data
    with open(original_json_path, "r", encoding="utf-8") as original_file:
        original_data = json.load(original_file)

    # Load translated JSON data
    with open(translated_json_path, "r", encoding="utf-8") as translated_file:
        translated_data = json.load(translated_file)

    # Convert translations to a dictionary {count: translated_value}
    translations = {str(item["count_src"]): item["translated"] for item in translated_data}

    # Track sheet name translations to apply at the end
    sheet_name_translations = {}

    # First pass: Collect sheet name translations
    for cell_info in original_data:
        count = str(cell_info["count_src"])  # Ensure count is a string
        if cell_info.get("is_sheet_name", False):
            original_sheet_name = cell_info["value"]
            translated_sheet_name = translations.get(count)
            if translated_sheet_name:
                if bilingual_mode:
                    bilingual_name = _format_bilingual_text(
                        original_sheet_name, translated_sheet_name, "sheet_name"
                    )
                    sheet_name_translations[original_sheet_name] = sanitize_sheet_name(bilingual_name)
                else:
                    sheet_name_translations[original_sheet_name] = translated_sheet_name.replace("␊", "\n").replace("␍", "\r")

    # Second pass: Update cell contents
    for cell_info in original_data:
        # Skip sheet name entries as they are handled separately
        if cell_info.get("is_sheet_name", False):
            continue

        count = str(cell_info["count_src"])  # Ensure count is a string
        sheet_name = cell_info["sheet"]
        row = cell_info["row"]
        column = cell_info["column"]
        original_text = cell_info["value"]
        is_merged = cell_info.get("is_merged", False)

        # Get the translated text
        translated_value = translations.get(count, None)
        if translated_value is None:
            # Log missing translation with original text
            app_logger.warning(
                f"Translation missing for count {count}. Original text: '{original_text}'"
            )
            continue

        # Format value based on mode
        if bilingual_mode:
            value = _format_bilingual_text(original_text, translated_value, "cell")
        else:
            value = translated_value.replace("␊", "\n").replace("␍", "\r")

        # Write to the Excel cell
        sheet = workbook[sheet_name]
        cell = sheet.cell(row=row, column=column)
        cell.value = value

        # Handle merged cells if applicable
        if is_merged:
            merge_range = f"{cell.coordinate}:{cell.coordinate}"
            sheet.merge_cells(merge_range)

    # Final pass: Rename sheets with their translations
    for original_name, translated_name in sheet_name_translations.items():
        if original_name in workbook.sheetnames:
            sheet = workbook[original_name]

            # Sanitize the sheet name to avoid invalid characters
            sanitized_name = sanitize_sheet_name(translated_name)

            # Log if the name needed to be changed
            if sanitized_name != translated_name:
                app_logger.warning(f"Sheet name '{translated_name}' contains invalid characters and was changed to '{sanitized_name}'")

            try:
                sheet.title = sanitized_name
                app_logger.info(f"Renamed sheet from '{original_name}' to '{sanitized_name}'")
            except Exception as e:
                app_logger.error(f"Failed to rename sheet from '{original_name}' to '{sanitized_name}': {str(e)}")
                # Keep the original name if renaming fails
                app_logger.info(f"Keeping original sheet name '{original_name}'")

    # Save the modified Excel file
    result_folder = os.path.join(result_dir)
    os.makedirs(result_folder, exist_ok=True)

    # Use source_lang2target_lang format if available, otherwise fallback to _translated
    if src_lang and dst_lang:
        lang_suffix = f"{src_lang}2{dst_lang}"
    else:
        lang_suffix = "translated"

    result_path = os.path.join(
        result_folder,
        f"{os.path.splitext(os.path.basename(file_path))[0]}_{lang_suffix}{os.path.splitext(file_path)[1]}"
    )

    try:
        workbook.save(result_path)
        app_logger.info(f"Translated Excel saved to: {result_path}")
    except Exception as e:
        app_logger.error(f"Failed to save translated Excel: {str(e)}")
        # Try saving with a different filename if there was an error
        fallback_path = os.path.join(
            result_folder,
            f"{os.path.splitext(os.path.basename(file_path))[0]}_{lang_suffix}_fallback{os.path.splitext(file_path)[1]}"
        )
        try:
            workbook.save(fallback_path)
            app_logger.info(f"Translated Excel saved to fallback path: {fallback_path}")
            result_path = fallback_path
        except Exception as e2:
            app_logger.error(f"Failed to save translated Excel to fallback path: {str(e2)}")
            raise

    return result_path


# ============================================================================
# XLWINGS MODE - EXTRACTION
# ============================================================================

def _extract_with_xlwings(file_path, temp_dir):
    """Extract Excel content using xlwings library (requires Excel installed)."""
    import xlwings as xw

    cell_data = []
    count = 0

    app = xw.App(visible=False)
    app.screen_updating = False

    try:
        wb = app.books.open(file_path)

        # First, collect sheet names
        sheets = list(wb.sheets)
        for sheet in sheets:
            sheet_name = sheet.name
            if should_translate(sheet_name):
                count += 1
                cell_data.append({
                    "count_src": count,
                    "sheet": sheet_name,
                    "value": sheet_name,
                    "type": "sheet_name"
                })

        def get_merged_ranges(sheet):
            """Get all merged cell ranges in the sheet."""
            merged_ranges = []
            try:
                merge_areas = sheet.api.Cells.MergeArea
                if merge_areas:
                    if hasattr(merge_areas, 'Address'):
                        merged_ranges.append(merge_areas.Address)
                    else:
                        for area in merge_areas:
                            if hasattr(area, 'Address'):
                                merged_ranges.append(area.Address)
            except:
                try:
                    used_range = sheet.used_range
                    if used_range:
                        max_row = used_range.last_cell.row
                        max_col = used_range.last_cell.column

                        for row in range(1, min(max_row + 1, 1000)):
                            for col in range(1, min(max_col + 1, 100)):
                                try:
                                    cell = sheet.cells(row, col)
                                    if cell.api.MergeCells:
                                        merge_area = cell.api.MergeArea
                                        if (merge_area.Row == row and merge_area.Column == col):
                                            address = merge_area.Address
                                            if address not in merged_ranges:
                                                merged_ranges.append(address)
                                except:
                                    continue
                except Exception as e:
                    app_logger.warning(f"Failed to get merged ranges for sheet {sheet.name}: {str(e)}")

            app_logger.info(f"Found {len(merged_ranges)} merged ranges in sheet '{sheet.name}'")
            return merged_ranges

        def parse_range_address(address):
            """Parse Excel address range, return (start_row, start_col, end_row, end_col)."""
            try:
                if '!' in address:
                    address = address.split('!')[-1]

                address = address.replace('$', '')

                if ':' in address:
                    start_addr, end_addr = address.split(':')
                else:
                    start_addr = end_addr = address

                def addr_to_row_col(addr):
                    col_str = ''.join([c for c in addr if c.isalpha()])
                    row_str = ''.join([c for c in addr if c.isdigit()])

                    col_num = 0
                    for c in col_str:
                        col_num = col_num * 26 + (ord(c.upper()) - ord('A') + 1)

                    return int(row_str), col_num

                start_row, start_col = addr_to_row_col(start_addr)
                end_row, end_col = addr_to_row_col(end_addr)

                return start_row, start_col, end_row, end_col
            except Exception as e:
                app_logger.warning(f"Failed to parse range address '{address}': {str(e)}")
                return None

        def is_cell_in_merged_ranges(row, col, merged_ranges, parsed_ranges_cache):
            """Check if cell is in a merged range and if it's the top-left cell."""
            for i, range_addr in enumerate(merged_ranges):
                if range_addr not in parsed_ranges_cache:
                    parsed = parse_range_address(range_addr)
                    if parsed:
                        parsed_ranges_cache[range_addr] = parsed
                    else:
                        continue

                start_row, start_col, end_row, end_col = parsed_ranges_cache[range_addr]

                if start_row <= row <= end_row and start_col <= col <= end_col:
                    is_top_left = (row == start_row and col == start_col)
                    return True, is_top_left, (start_row, start_col, end_row, end_col)

            return False, False, None

        def process_sheet(sheet):
            nonlocal count
            sheet_data = []

            merged_ranges = get_merged_ranges(sheet)
            parsed_ranges_cache = {}

            try:
                used_range = sheet.used_range
                if used_range:
                    app_logger.info(f"Processing sheet '{sheet.name}' with range: {used_range.address}")

                    max_row = used_range.last_cell.row
                    max_col = used_range.last_cell.column

                    app_logger.info(f"Sheet '{sheet.name}' dimensions: {max_row} rows x {max_col} columns")

                    batch_size = 5000
                    processed_cells = 0

                    for batch_start in range(1, max_row + 1, batch_size):
                        batch_end = min(batch_start + batch_size - 1, max_row)
                        app_logger.info(f"Processing rows {batch_start} to {batch_end}")

                        for row_idx in range(batch_start, batch_end + 1):
                            for col_idx in range(1, max_col + 1):
                                try:
                                    cell = sheet.cells(row_idx, col_idx)
                                    cell_value = cell.value

                                    if cell_value is None:
                                        continue

                                    if isinstance(cell_value, datetime):
                                        continue

                                    cell_value_str = str(cell_value)

                                    if cell_value_str.strip().startswith('='):
                                        continue

                                    if not should_translate(cell_value_str):
                                        continue

                                    is_merged, is_top_left, merge_info = is_cell_in_merged_ranges(
                                        row_idx, col_idx, merged_ranges, parsed_ranges_cache
                                    )

                                    if is_merged and not is_top_left:
                                        continue

                                    processed_value = cell_value_str.replace("\n", "␊").replace("\r", "␍")

                                    count += 1
                                    cell_info = {
                                        "count_src": count,
                                        "sheet": sheet.name,
                                        "row": row_idx,
                                        "column": col_idx,
                                        "value": processed_value,
                                        "original_value": processed_value,
                                        "is_merged": is_merged,
                                        "type": "cell"
                                    }

                                    if is_merged and merge_info:
                                        cell_info["merge_start_row"] = merge_info[0]
                                        cell_info["merge_start_col"] = merge_info[1]
                                        cell_info["merge_end_row"] = merge_info[2]
                                        cell_info["merge_end_col"] = merge_info[3]

                                    sheet_data.append(cell_info)
                                    processed_cells += 1

                                    if processed_cells % 1000 == 0:
                                        app_logger.info(f"Processed {processed_cells} cells in sheet '{sheet.name}'")

                                except Exception as cell_error:
                                    app_logger.warning(f"Error processing cell ({row_idx}, {col_idx}): {str(cell_error)}")
                                    continue

                    app_logger.info(f"Extracted {len(sheet_data)} cells from sheet '{sheet.name}'")

            except Exception as e:
                app_logger.error(f"Error processing cells in sheet {sheet.name}: {str(e)}")

            # Process shapes
            try:
                shapes = list(sheet.shapes)
                if shapes:
                    app_logger.info(f"Processing {len(shapes)} shapes in sheet '{sheet.name}'")
                    shape_name_count = {}

                    def process_group_items(group, group_index, group_name, path=""):
                        group_items_data = []

                        try:
                            if hasattr(group.api, 'GroupItems'):
                                group_items = group.api.GroupItems
                                for i in range(1, group_items.Count + 1):
                                    try:
                                        child_item = group_items.Item(i)
                                        item_path = f"{path}/{i}" if path else str(i)

                                        is_child_group = False
                                        try:
                                            if hasattr(child_item, 'Type') and child_item.Type == 6:
                                                is_child_group = True
                                        except:
                                            pass

                                        if is_child_group:
                                            try:
                                                child_name = f"{group_name}_child{i}"
                                                nested_items = process_group_items(child_item, -1, child_name, item_path)
                                                group_items_data.extend(nested_items)
                                            except Exception as nested_error:
                                                app_logger.warning(f"Error processing nested group {item_path}: {str(nested_error)}")
                                        else:
                                            has_text = False
                                            text_value = None

                                            try:
                                                if hasattr(child_item, 'TextFrame') and child_item.TextFrame.HasText:
                                                    text_value = child_item.TextFrame.Characters().Text
                                                    has_text = True
                                            except:
                                                pass

                                            if not has_text:
                                                try:
                                                    if hasattr(child_item, 'TextFrame2') and child_item.TextFrame2.HasText:
                                                        text_value = child_item.TextFrame2.TextRange.Text
                                                        has_text = True
                                                except:
                                                    pass

                                            if has_text and text_value and isinstance(text_value, str) and text_value.strip().startswith('='):
                                                continue

                                            if has_text and text_value and should_translate(text_value):
                                                text_value = str(text_value).replace("\n", "␊").replace("\r", "␍")

                                                try:
                                                    child_name = child_item.Name
                                                except:
                                                    child_name = f"GroupChild_{group_name}_{item_path}"

                                                if child_name in shape_name_count:
                                                    shape_name_count[child_name] += 1
                                                else:
                                                    shape_name_count[child_name] = 1

                                                unique_shape_id = f"{child_name}_{shape_name_count[child_name]}"

                                                count += 1
                                                group_items_data.append({
                                                    "count_src": count,
                                                    "sheet": sheet.name,
                                                    "shape_name": child_name,
                                                    "unique_shape_id": unique_shape_id,
                                                    "shape_index": -1,
                                                    "group_name": group_name,
                                                    "group_index": group_index,
                                                    "child_path": item_path,
                                                    "value": text_value,
                                                    "original_value": text_value,
                                                    "type": "group_textbox"
                                                })
                                    except Exception as child_error:
                                        app_logger.warning(f"Error processing group child {path}/{i}: {str(child_error)}")
                                        continue
                        except Exception as group_error:
                            app_logger.warning(f"Error accessing group items: {str(group_error)}")

                        return group_items_data

                    for shape_idx, shape in enumerate(shapes):
                        try:
                            is_group = False
                            try:
                                if hasattr(shape, 'type') and 'group' in str(shape.type).lower():
                                    is_group = True
                            except:
                                try:
                                    if shape.api.Type == 6:
                                        is_group = True
                                except:
                                    pass

                            if is_group:
                                try:
                                    group_name = shape.name
                                    group_items_data = process_group_items(shape, shape_idx, group_name)
                                    sheet_data.extend(group_items_data)
                                except Exception as group_error:
                                    app_logger.warning(f"Error processing group {shape_idx}: {str(group_error)}")
                            else:
                                try:
                                    if hasattr(shape, 'text') and shape.text:
                                        text_value = shape.text

                                        if isinstance(text_value, str) and text_value.strip().startswith('='):
                                            continue

                                        if not should_translate(text_value):
                                            continue

                                        text_value = str(text_value).replace("\n", "␊").replace("\r", "␍")

                                        original_shape_name = shape.name
                                        if original_shape_name in shape_name_count:
                                            shape_name_count[original_shape_name] += 1
                                        else:
                                            shape_name_count[original_shape_name] = 1
                                        unique_shape_id = f"{original_shape_name}_{shape_name_count[original_shape_name]}"

                                        count += 1
                                        sheet_data.append({
                                            "count_src": count,
                                            "sheet": sheet.name,
                                            "shape_name": original_shape_name,
                                            "unique_shape_id": unique_shape_id,
                                            "shape_index": shape_idx,
                                            "value": text_value,
                                            "original_value": text_value,
                                            "type": "textbox"
                                        })
                                except Exception as shape_error:
                                    app_logger.warning(f"Error processing individual shape {shape_idx}: {str(shape_error)}")
                        except Exception as e:
                            app_logger.warning(f"Error processing shape #{shape_idx}: {str(e)}")
                            continue

                    app_logger.info(f"Extracted {len([item for item in sheet_data if item['type'] in ['textbox', 'group_textbox']])} textboxes from sheet '{sheet.name}'")

            except Exception as e:
                app_logger.error(f"Error processing shapes in sheet {sheet.name}: {str(e)}")

            return sheet_data

        # Process all sheets
        for sheet in sheets:
            try:
                sheet_data = process_sheet(sheet)
                cell_data.extend(sheet_data)
            except Exception as e:
                app_logger.error(f"Error processing sheet {sheet.name}: {str(e)}")

        app_logger.info(f"Extracted {len(cell_data)} items using xlwings")

    finally:
        try:
            wb.close()
            app.quit()
        except Exception as e:
            app_logger.error(f"Error closing workbook: {str(e)}")

    # Extract SmartArt content using ZIP operations
    try:
        count = _extract_smartart_from_excel(file_path, cell_data, count)
    except Exception as e:
        app_logger.error(f"Error extracting SmartArt from Excel: {str(e)}")

    # Extract drawing content using ZIP operations
    try:
        count = _extract_drawing_content_from_excel(file_path, cell_data, count)
    except Exception as e:
        app_logger.error(f"Error extracting drawing content from Excel: {str(e)}")

    # Save to JSON
    filename = os.path.splitext(os.path.basename(file_path))[0]
    temp_folder = os.path.join(temp_dir, filename)
    os.makedirs(temp_folder, exist_ok=True)
    json_path = os.path.join(temp_folder, "src.json")

    with open(json_path, "w", encoding="utf-8") as json_file:
        json.dump(cell_data, json_file, ensure_ascii=False, indent=4)

    app_logger.info(f"Total extracted items: {len(cell_data)}")
    return json_path


# ============================================================================
# XLWINGS MODE - EXTRACTION HELPERS (SmartArt and Drawing)
# ============================================================================

def _extract_drawing_content_from_excel(file_path: str, content_data: List, count: int) -> int:
    """Extract text from drawing files in Excel."""
    try:
        namespaces = {
            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        }

        with ZipFile(file_path, 'r') as excel_zip:
            drawing_files = [name for name in excel_zip.namelist()
                            if name.startswith('xl/drawings/drawing') and name.endswith('.xml')]
            drawing_files.sort()

            app_logger.info(f"Found {len(drawing_files)} drawing files in Excel")

            sheet_drawing_map = _get_sheet_drawing_map(excel_zip)

            for drawing_path in drawing_files:
                try:
                    drawing_match = re.search(r'drawing(\d+)\.xml', drawing_path)
                    if not drawing_match:
                        continue

                    drawing_index = int(drawing_match.group(1))
                    sheet_name = sheet_drawing_map.get(drawing_index, f"Sheet{drawing_index}")

                    drawing_xml = excel_zip.read(drawing_path)
                    drawing_tree = etree.fromstring(drawing_xml)

                    count = _extract_textboxes_from_drawing(
                        drawing_tree, namespaces, content_data, count,
                        drawing_path, sheet_name, drawing_index
                    )

                except Exception as e:
                    app_logger.error(f"Failed to extract drawing content from {drawing_path}: {e}")
                    continue

    except Exception as e:
        app_logger.error(f"Failed to extract drawing content from Excel: {e}")

    drawing_items = sum(1 for item in content_data if item.get('type') == 'excel_drawing')
    app_logger.info(f"Extracted {drawing_items} drawing textbox items from Excel")
    return count


def _get_sheet_drawing_map(excel_zip) -> Dict[int, str]:
    """Get mapping of drawing index to sheet name."""
    sheet_drawing_map = {}

    try:
        if 'xl/workbook.xml' in excel_zip.namelist():
            workbook_xml = excel_zip.read('xl/workbook.xml')
            workbook_tree = etree.fromstring(workbook_xml)

            sheets = workbook_tree.xpath('.//sheet', namespaces={'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'})

            for sheet in sheets:
                sheet_name = sheet.get('name')
                sheet_id = sheet.get('sheetId')

                worksheet_path = f'xl/worksheets/sheet{sheet_id}.xml'
                if worksheet_path in excel_zip.namelist():
                    try:
                        worksheet_xml = excel_zip.read(worksheet_path)
                        worksheet_tree = etree.fromstring(worksheet_xml)

                        drawings = worksheet_tree.xpath('.//drawing', namespaces={'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'})
                        for drawing in drawings:
                            drawing_r_id = drawing.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                            if drawing_r_id:
                                if sheet_id:
                                    sheet_drawing_map[int(sheet_id)] = sheet_name

                    except Exception as e:
                        app_logger.warning(f"Error processing worksheet {worksheet_path}: {e}")

    except Exception as e:
        app_logger.warning(f"Error getting sheet-drawing mapping: {e}")

    return sheet_drawing_map


def _extract_textboxes_from_drawing(drawing_tree, namespaces: Dict, content_data: List,
                                   count: int, drawing_path: str, sheet_name: str,
                                   drawing_index: int) -> int:
    """Extract text from all textboxes in a drawing."""
    anchors = drawing_tree.xpath('.//xdr:twoCellAnchor', namespaces=namespaces)

    for anchor_idx, anchor in enumerate(anchors):
        count = _process_drawing_anchor(
            anchor, namespaces, content_data, count,
            drawing_path, sheet_name, drawing_index, anchor_idx
        )

    return count


def _process_drawing_anchor(anchor, namespaces: Dict, content_data: List, count: int,
                           drawing_path: str, sheet_name: str, drawing_index: int,
                           anchor_idx: int) -> int:
    """Process a single anchor element and extract all text content."""
    textboxes = anchor.xpath('.//xdr:sp[.//xdr:txBody]', namespaces=namespaces)
    for tb_idx, textbox in enumerate(textboxes):
        count = _process_drawing_textbox(
            textbox, namespaces, content_data, count, drawing_path,
            sheet_name, drawing_index, anchor_idx, tb_idx, "textbox"
        )

    groups = anchor.xpath('.//xdr:grpSp', namespaces=namespaces)
    for group_idx, group in enumerate(groups):
        count = _process_drawing_group(
            group, namespaces, content_data, count, drawing_path,
            sheet_name, drawing_index, anchor_idx, group_idx, ""
        )

    return count


def _process_drawing_group(group, namespaces: Dict, content_data: List, count: int,
                          drawing_path: str, sheet_name: str, drawing_index: int,
                          anchor_idx: int, group_idx: int, group_path: str) -> int:
    """Process a group shape and extract text from all nested elements."""
    current_path = f"{group_path}/G{group_idx}" if group_path else f"G{group_idx}"

    group_name = ""
    try:
        cNvPr = group.xpath('.//xdr:cNvPr', namespaces=namespaces)
        if cNvPr:
            group_name = cNvPr[0].get('name', f"Group_{group_idx}")
    except:
        group_name = f"Group_{group_idx}"

    textboxes = group.xpath('./xdr:sp[.//xdr:txBody]', namespaces=namespaces)
    for tb_idx, textbox in enumerate(textboxes):
        count = _process_drawing_textbox(
            textbox, namespaces, content_data, count, drawing_path,
            sheet_name, drawing_index, anchor_idx, tb_idx, "group_textbox",
            group_name, current_path
        )

    nested_groups = group.xpath('./xdr:grpSp', namespaces=namespaces)
    for nested_idx, nested_group in enumerate(nested_groups):
        count = _process_drawing_group(
            nested_group, namespaces, content_data, count, drawing_path,
            sheet_name, drawing_index, anchor_idx, nested_idx, current_path
        )

    return count


def _process_drawing_textbox(textbox, namespaces: Dict, content_data: List, count: int,
                            drawing_path: str, sheet_name: str, drawing_index: int,
                            anchor_idx: int, tb_idx: int, textbox_type: str,
                            group_name: str = "", group_path: str = "") -> int:
    """Process a single textbox and extract its text content."""
    try:
        textbox_name = ""
        textbox_id = ""
        try:
            cNvPr = textbox.xpath('.//xdr:cNvPr', namespaces=namespaces)
            if cNvPr:
                textbox_name = cNvPr[0].get('name', f"TextBox_{tb_idx}")
                textbox_id = cNvPr[0].get('id', str(tb_idx))
        except:
            textbox_name = f"TextBox_{tb_idx}"
            textbox_id = str(tb_idx)

        tx_bodies = textbox.xpath('.//xdr:txBody', namespaces=namespaces)

        for tx_body_idx, tx_body in enumerate(tx_bodies):
            paragraphs = tx_body.xpath('.//a:p', namespaces=namespaces)

            for p_idx, paragraph in enumerate(paragraphs):
                text_runs = paragraph.xpath('.//a:r', namespaces=namespaces)

                if not text_runs:
                    text_nodes = paragraph.xpath('.//a:t', namespaces=namespaces)
                    if text_nodes and text_nodes[0].text:
                        merged_text = text_nodes[0].text.strip()
                        if merged_text and should_translate(merged_text):
                            count += 1
                            content_data.append({
                                "count_src": count,
                                "sheet": sheet_name,
                                "drawing_path": drawing_path,
                                "drawing_index": drawing_index,
                                "anchor_index": anchor_idx,
                                "textbox_index": tb_idx,
                                "tx_body_index": tx_body_idx,
                                "paragraph_index": p_idx,
                                "textbox_name": textbox_name,
                                "textbox_id": textbox_id,
                                "group_name": group_name,
                                "group_path": group_path,
                                "type": "excel_drawing",
                                "textbox_type": textbox_type,
                                "value": merged_text.replace("\n", "␊").replace("\r", "␍"),
                                "original_value": merged_text.replace("\n", "␊").replace("\r", "␍"),
                                "original_text": merged_text,
                                "run_texts": [merged_text],
                                "run_styles": [{}],
                                "run_lengths": [len(merged_text)],
                                "xpath": f".//xdr:twoCellAnchor[{anchor_idx + 1}]//xdr:sp[{tb_idx + 1}]//xdr:txBody[{tx_body_idx + 1}]//a:p[{p_idx + 1}]"
                            })
                    continue

                run_info = _process_drawing_text_runs(text_runs, namespaces)

                if not run_info['merged_text'].strip():
                    continue

                if should_translate(run_info['merged_text']):
                    count += 1
                    content_data.append({
                        "count_src": count,
                        "sheet": sheet_name,
                        "drawing_path": drawing_path,
                        "drawing_index": drawing_index,
                        "anchor_index": anchor_idx,
                        "textbox_index": tb_idx,
                        "tx_body_index": tx_body_idx,
                        "paragraph_index": p_idx,
                        "textbox_name": textbox_name,
                        "textbox_id": textbox_id,
                        "group_name": group_name,
                        "group_path": group_path,
                        "type": "excel_drawing",
                        "textbox_type": textbox_type,
                        "value": run_info['merged_text'].replace("\n", "␊").replace("\r", "␍"),
                        "original_value": run_info['merged_text'].replace("\n", "␊").replace("\r", "␍"),
                        "run_texts": run_info['run_texts'],
                        "run_styles": run_info['run_styles'],
                        "run_lengths": run_info['run_lengths'],
                        "original_text": run_info['merged_text'],
                        "xpath": f".//xdr:twoCellAnchor[{anchor_idx + 1}]//xdr:sp[{tb_idx + 1}]//xdr:txBody[{tx_body_idx + 1}]//a:p[{p_idx + 1}]"
                    })

    except Exception as e:
        app_logger.warning(f"Error processing drawing textbox: {e}")

    return count


def _process_drawing_text_runs(text_runs, namespaces: dict) -> dict:
    """Process text runs for Excel drawing."""
    merged_text = ""
    run_texts = []
    run_styles = []
    run_lengths = []

    for text_run in text_runs:
        text_node = text_run.xpath('./a:t', namespaces=namespaces)
        if text_node and text_node[0].text is not None:
            run_text = text_node[0].text
        else:
            run_text = ""

        merged_text += run_text
        run_texts.append(run_text)
        run_lengths.append(len(run_text))
        run_styles.append(_extract_drawing_run_style(text_run, namespaces))

    return {
        'merged_text': merged_text,
        'run_texts': run_texts,
        'run_styles': run_styles,
        'run_lengths': run_lengths
    }


def _extract_drawing_run_style(text_run, namespaces: dict) -> dict:
    """Extract style information from a text run in Excel drawing."""
    style_info = {}

    try:
        rpr = text_run.xpath('./a:rPr', namespaces=namespaces)
        if rpr:
            rpr_element = rpr[0]

            sz = rpr_element.get('sz')
            if sz:
                style_info['font_size'] = sz

            b = rpr_element.get('b')
            if b:
                style_info['bold'] = b

            i = rpr_element.get('i')
            if i:
                style_info['italic'] = i

            u = rpr_element.get('u')
            if u:
                style_info['underline'] = u

            latin = rpr_element.xpath('./a:latin', namespaces=namespaces)
            if latin:
                style_info['font_family'] = latin[0].get('typeface')

            solid_fill = rpr_element.xpath('./a:solidFill/a:srgbClr', namespaces=namespaces)
            if solid_fill:
                style_info['color'] = solid_fill[0].get('val')

            strike = rpr_element.get('strike')
            if strike:
                style_info['strike'] = strike

    except Exception as e:
        app_logger.warning(f"Failed to extract drawing style information: {e}")

    return style_info


def _extract_smartart_from_excel(file_path: str, content_data: List, count: int) -> int:
    """Extract text from SmartArt diagrams in Excel."""
    try:
        namespaces = {
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'dgm': 'http://schemas.openxmlformats.org/drawingml/2006/diagram',
            'dsp': 'http://schemas.microsoft.com/office/drawing/2008/diagram',
            'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        }

        with ZipFile(file_path, 'r') as excel_zip:
            diagram_drawings = [name for name in excel_zip.namelist()
                               if name.startswith('xl/diagrams/drawing') and name.endswith('.xml')]
            diagram_drawings.sort()

            app_logger.info(f"Found {len(diagram_drawings)} SmartArt diagram files in Excel")

            for drawing_path in diagram_drawings:
                try:
                    diagram_match = re.search(r'drawing(\d+)\.xml', drawing_path)
                    if not diagram_match:
                        continue

                    diagram_index = int(diagram_match.group(1))

                    drawing_xml = excel_zip.read(drawing_path)
                    drawing_tree = etree.fromstring(drawing_xml)

                    shapes = drawing_tree.xpath('.//dsp:sp[.//dsp:txBody]', namespaces=namespaces)

                    for shape_index, shape in enumerate(shapes):
                        model_id = shape.get('modelId', '')

                        tx_bodies = shape.xpath('.//dsp:txBody', namespaces=namespaces)

                        for tx_body_index, tx_body in enumerate(tx_bodies):
                            paragraphs = tx_body.xpath('.//a:p', namespaces=namespaces)

                            for p_index, paragraph in enumerate(paragraphs):
                                text_runs = paragraph.xpath('.//a:r', namespaces=namespaces)

                                if not text_runs:
                                    continue

                                run_info = _process_excel_smartart_text_runs(text_runs, namespaces)

                                if not run_info['merged_text'].strip():
                                    continue

                                if should_translate(run_info['merged_text']):
                                    count += 1
                                    content_data.append({
                                        "count_src": count,
                                        "diagram_index": diagram_index,
                                        "shape_index": shape_index,
                                        "tx_body_index": tx_body_index,
                                        "paragraph_index": p_index,
                                        "model_id": model_id,
                                        "type": "excel_smartart",
                                        "value": run_info['merged_text'].replace("\n", "␊").replace("\r", "␍"),
                                        "original_value": run_info['merged_text'].replace("\n", "␊").replace("\r", "␍"),
                                        "run_texts": run_info['run_texts'],
                                        "run_styles": run_info['run_styles'],
                                        "run_lengths": run_info['run_lengths'],
                                        "drawing_path": drawing_path,
                                        "original_text": run_info['merged_text'],
                                        "xpath": f".//dsp:sp[{shape_index + 1}]//dsp:txBody[{tx_body_index + 1}]//a:p[{p_index + 1}]"
                                    })

                except Exception as e:
                    app_logger.error(f"Failed to extract SmartArt from {drawing_path}: {e}")
                    continue

    except Exception as e:
        app_logger.error(f"Failed to extract SmartArt from Excel: {e}")

    app_logger.info(f"Extracted {sum(1 for item in content_data if item.get('type') == 'excel_smartart')} SmartArt items from Excel")
    return count


def _process_excel_smartart_text_runs(text_runs, namespaces: dict) -> dict:
    """Process text runs for Excel SmartArt."""
    merged_text = ""
    run_texts = []
    run_styles = []
    run_lengths = []

    for text_run in text_runs:
        text_node = text_run.xpath('./a:t', namespaces=namespaces)
        if text_node and text_node[0].text is not None:
            run_text = text_node[0].text
        else:
            run_text = ""

        merged_text += run_text
        run_texts.append(run_text)
        run_lengths.append(len(run_text))
        run_styles.append(_extract_excel_smartart_run_style(text_run, namespaces))

    return {
        'merged_text': merged_text,
        'run_texts': run_texts,
        'run_styles': run_styles,
        'run_lengths': run_lengths
    }


def _extract_excel_smartart_run_style(text_run, namespaces: dict) -> dict:
    """Extract style information from a text run in Excel SmartArt."""
    style_info = {}

    try:
        rpr = text_run.xpath('./a:rPr', namespaces=namespaces)
        if rpr:
            rpr_element = rpr[0]

            sz = rpr_element.get('sz')
            if sz:
                style_info['font_size'] = sz

            b = rpr_element.get('b')
            if b:
                style_info['bold'] = b

            i = rpr_element.get('i')
            if i:
                style_info['italic'] = i

            u = rpr_element.get('u')
            if u:
                style_info['underline'] = u

            latin = rpr_element.xpath('./a:latin', namespaces=namespaces)
            if latin:
                style_info['font_family'] = latin[0].get('typeface')

            solid_fill = rpr_element.xpath('./a:solidFill/a:srgbClr', namespaces=namespaces)
            if solid_fill:
                style_info['color'] = solid_fill[0].get('val')

            strike = rpr_element.get('strike')
            if strike:
                style_info['strike'] = strike

    except Exception as e:
        app_logger.warning(f"Failed to extract Excel SmartArt style information: {e}")

    return style_info


# ============================================================================
# XLWINGS MODE - WRITING
# ============================================================================

def _write_with_xlwings(file_path, original_json_path, translated_json_path, result_dir, bilingual_mode=False, src_lang=None, dst_lang=None):
    """Write translated content using xlwings library."""
    import xlwings as xw

    # Load JSON data
    with open(original_json_path, "r", encoding="utf-8") as original_file:
        original_data = json.load(original_file)

    with open(translated_json_path, "r", encoding="utf-8") as translated_file:
        translated_data = json.load(translated_file)

    app_logger.info(f"Original data count: {len(original_data)}")
    app_logger.info(f"Translation data count: {len(translated_data)}")

    translations = {}
    for item in translated_data:
        count_src = str(item["count_src"])
        translations[count_src] = item["translated"]

    app_logger.info(f"Translation dictionary count: {len(translations)}")

    missing_translations = []
    original_count_srcs = set(str(item["count_src"]) for item in original_data)
    translated_count_srcs = set(translations.keys())

    missing_in_translation = original_count_srcs - translated_count_srcs
    extra_in_translation = translated_count_srcs - original_count_srcs

    if missing_in_translation:
        app_logger.warning(f"Missing translations for count_src: {sorted(missing_in_translation)}")
    if extra_in_translation:
        app_logger.warning(f"Extra translations for count_src: {sorted(extra_in_translation)}")

    # Collect sheet name translations
    sheet_name_translations = {}
    for cell_info in original_data:
        if cell_info.get("type") == "sheet_name":
            count_src = str(cell_info["count_src"])
            original_sheet_name = cell_info["sheet"]
            translated_sheet_name = translations.get(count_src)
            if translated_sheet_name:
                if bilingual_mode:
                    # Format as bilingual sheet name
                    bilingual_name = _format_bilingual_text(
                        cell_info["value"], translated_sheet_name, "sheet_name"
                    )
                    sanitized_name = sanitize_sheet_name(bilingual_name)
                else:
                    sanitized_name = sanitize_sheet_name(
                        translated_sheet_name.replace("␊", "\n").replace("␍", "\r")
                    )
                sheet_name_translations[original_sheet_name] = sanitized_name

                if sanitized_name != translated_sheet_name.replace("␊", "\n").replace("␍", "\r"):
                    app_logger.warning(f"Sheet name '{translated_sheet_name}' was changed to '{sanitized_name}'")

    # Organize data by sheet and type
    sheets_data = {}
    smartart_items = []
    drawing_items = []

    for cell_info in original_data:
        if cell_info.get("type") == "sheet_name":
            continue
        elif cell_info.get("type") == "excel_smartart":
            smartart_items.append(cell_info)
            continue
        elif cell_info.get("type") == "excel_drawing":
            drawing_items.append(cell_info)
            continue

        count_src = str(cell_info["count_src"])
        sheet_name = cell_info["sheet"]

        if sheet_name not in sheets_data:
            sheets_data[sheet_name] = {
                "cells": [],
                "textboxes": []
            }

        translated_value = translations.get(count_src)
        if translated_value is None:
            missing_translations.append(count_src)
            app_logger.warning(
                f"Translation missing for count_src {count_src}. Sheet: {sheet_name}, Original text: '{cell_info['value'][:100]}...'"
            )
            translated_value = cell_info['value']

        # Format value based on mode
        if bilingual_mode:
            final_value = _format_bilingual_text(
                cell_info['value'],
                translated_value,
                "cell" if cell_info.get("type") == "cell" else "textbox"
            )
        else:
            final_value = translated_value.replace("␊", "\n").replace("␍", "\r")

        if cell_info.get("type") == "cell":
            cell_data = {
                "row": cell_info["row"],
                "column": cell_info["column"],
                "value": final_value,
                "is_merged": cell_info.get("is_merged", False),
                "count_src": count_src,
                "original_value": cell_info.get("original_value", cell_info["value"])
            }

            if cell_info.get("is_merged"):
                cell_data["merge_start_row"] = cell_info.get("merge_start_row")
                cell_data["merge_start_col"] = cell_info.get("merge_start_col")
                cell_data["merge_end_row"] = cell_info.get("merge_end_row")
                cell_data["merge_end_col"] = cell_info.get("merge_end_col")

            sheets_data[sheet_name]["cells"].append(cell_data)
        else:
            textbox_data = cell_info.copy()
            textbox_data["value"] = final_value
            textbox_data["original_value"] = cell_info.get("original_value", cell_info["value"])
            sheets_data[sheet_name]["textboxes"].append(textbox_data)

    if missing_translations:
        app_logger.warning(f"Found {len(missing_translations)} missing translations, using original text")

    # Open Excel application
    app = xw.App(visible=False)
    app.screen_updating = False
    app.display_alerts = False

    try:
        wb = app.books.open(file_path)

        # Handle sheet renaming
        original_to_translated_sheet_map = {}
        new_sheet_names = []

        for sheet_name, data in sheets_data.items():
            translated_sheet_name = sheet_name_translations.get(sheet_name)
            if translated_sheet_name:
                original_to_translated_sheet_map[sheet_name] = translated_sheet_name
                new_sheet_names.append((sheet_name, translated_sheet_name))

        existing_names = set(sheet.name for sheet in wb.sheets)
        temp_names = {}

        # First pass: Use temporary names to avoid conflicts
        for original_name, new_name in new_sheet_names:
            if new_name in existing_names and new_name != original_name:
                temp_name = f"temp_{original_name}_{hash(original_name) % 10000}"
                temp_names[original_name] = temp_name

        for original_name, temp_name in temp_names.items():
            try:
                wb.sheets[original_name].name = temp_name
                app_logger.info(f"Temporarily renamed sheet '{original_name}' to '{temp_name}'")
            except Exception as e:
                app_logger.warning(f"Error temporarily renaming sheet '{original_name}' to '{temp_name}': {str(e)}")

        # Second pass: Rename to final translated names
        for original_name, new_name in new_sheet_names:
            actual_original_name = temp_names.get(original_name, original_name)
            try:
                if actual_original_name == new_name:
                    app_logger.info(f"Sheet '{original_name}' translation is identical to original, skipping rename")
                    continue

                wb.sheets[actual_original_name].name = new_name
                app_logger.info(f"Successfully renamed sheet '{original_name}' to '{new_name}'")
            except Exception as e:
                app_logger.warning(f"Error renaming sheet '{original_name}' to '{new_name}': {str(e)}")
                if original_name in temp_names:
                    try:
                        wb.sheets[actual_original_name].name = original_name
                        app_logger.info(f"Restored original sheet name '{original_name}'")
                    except Exception as restore_err:
                        app_logger.error(f"Failed to restore original sheet name '{original_name}': {str(restore_err)}")

        # Update sheet data references to use translated names
        updated_sheets_data = {}
        for sheet_name, data in sheets_data.items():
            actual_sheet_name = sheet_name_translations.get(sheet_name, sheet_name)
            if actual_sheet_name in [sheet.name for sheet in wb.sheets]:
                updated_sheets_data[actual_sheet_name] = data
            else:
                updated_sheets_data[sheet_name] = data

        # Process cell and shape content
        for sheet_name, data in updated_sheets_data.items():
            try:
                sheet = wb.sheets[sheet_name]

                app_logger.info(f"Processing {len(data['cells'])} cells in sheet '{sheet_name}'")

                cells_to_process = sorted(data["cells"], key=lambda x: (x["row"], x["column"]))

                successful_updates = 0
                failed_updates = 0

                batch_size = 1000
                for batch_start in range(0, len(cells_to_process), batch_size):
                    batch_end = min(batch_start + batch_size, len(cells_to_process))
                    app_logger.info(f"Processing cell batch {batch_start//batch_size + 1}/{(len(cells_to_process)-1)//batch_size + 1}")

                    for cell_data in cells_to_process[batch_start:batch_end]:
                        try:
                            row = cell_data["row"]
                            column = cell_data["column"]
                            value = cell_data["value"]
                            is_merged = cell_data.get("is_merged", False)
                            count_src = cell_data.get("count_src", "unknown")

                            if row < 1 or column < 1:
                                app_logger.warning(f"Invalid cell position ({row}, {column}) for count_src {count_src}, skipping")
                                failed_updates += 1
                                continue

                            if is_merged:
                                merge_start_row = cell_data.get("merge_start_row", row)
                                merge_start_col = cell_data.get("merge_start_col", column)

                                if row != merge_start_row or column != merge_start_col:
                                    app_logger.warning(f"Cell ({row}, {column}) count_src {count_src} is merged but not at top-left position. Using ({merge_start_row}, {merge_start_col}) instead")
                                    row = merge_start_row
                                    column = merge_start_col

                            sheet.cells(row, column).value = value
                            successful_updates += 1

                            if successful_updates % 100 == 0:
                                app_logger.debug(f"Successfully updated {successful_updates} cells in sheet '{sheet_name}'")

                        except Exception as cell_error:
                            failed_updates += 1
                            app_logger.warning(f"Error updating cell ({row}, {column}) count_src {count_src}: {str(cell_error)}")
                            continue

                app_logger.info(f"Cell processing completed for sheet '{sheet_name}': {successful_updates} successful, {failed_updates} failed")

                # Process textboxes
                app_logger.info(f"Processing {len(data['textboxes'])} textboxes in sheet '{sheet_name}'")

                try:
                    all_shapes = list(sheet.shapes)
                except Exception as shapes_error:
                    app_logger.warning(f"Error getting shapes from sheet '{sheet_name}': {str(shapes_error)}")
                    all_shapes = []

                normal_textboxes = [tb for tb in data["textboxes"] if tb.get("type") == "textbox"]
                group_textboxes = [tb for tb in data["textboxes"] if tb.get("type") == "group_textbox"]

                textbox_success = 0
                textbox_failed = 0

                # Process normal textboxes
                for textbox in normal_textboxes:
                    try:
                        matched = False
                        shape_index = textbox.get("shape_index")
                        count_src = textbox.get("count_src", "unknown")

                        if shape_index is not None and 0 <= shape_index < len(all_shapes):
                            try:
                                shape = all_shapes[shape_index]
                                if hasattr(shape, 'text'):
                                    shape.text = textbox["value"]
                                    matched = True
                                    textbox_success += 1
                                    app_logger.debug(f"Updated textbox by index {shape_index}, count_src {count_src}")
                            except Exception as e:
                                app_logger.warning(f"Error updating shape by index {shape_index}, count_src {count_src}: {str(e)}")

                        if not matched and textbox.get("unique_shape_id"):
                            try:
                                original_name = textbox["shape_name"]
                                same_name_shapes = [s for s in all_shapes if s.name == original_name]
                                unique_id_parts = textbox["unique_shape_id"].split("_")
                                if len(unique_id_parts) > 1:
                                    id_number = int(unique_id_parts[-1])
                                    if 1 <= id_number <= len(same_name_shapes):
                                        shape = same_name_shapes[id_number - 1]
                                        if hasattr(shape, 'text'):
                                            shape.text = textbox["value"]
                                            matched = True
                                            textbox_success += 1
                                            app_logger.debug(f"Updated shape by unique ID: {textbox['unique_shape_id']}, count_src {count_src}")
                            except Exception as e:
                                app_logger.warning(f"Error updating shape with unique ID {textbox['unique_shape_id']}, count_src {count_src}: {str(e)}")

                        if not matched:
                            try:
                                for shape in all_shapes:
                                    if shape.name == textbox["shape_name"]:
                                        if hasattr(shape, 'text'):
                                            shape.text = textbox["value"]
                                            matched = True
                                            textbox_success += 1
                                            app_logger.debug(f"Updated shape by name: {textbox['shape_name']}, count_src {count_src}")
                                            break
                            except Exception as e:
                                app_logger.warning(f"Error updating shape by name {textbox['shape_name']}, count_src {count_src}: {str(e)}")

                        if not matched:
                            textbox_failed += 1
                            app_logger.warning(f"Could not find shape to update: {textbox.get('shape_name', 'unknown')}, count_src {count_src}")

                    except Exception as textbox_error:
                        textbox_failed += 1
                        app_logger.warning(f"Error processing textbox count_src {textbox.get('count_src', 'unknown')}: {str(textbox_error)}")
                        continue

                # Process group textboxes
                for textbox in group_textboxes:
                    try:
                        count_src = textbox.get("count_src", "unknown")
                        group_name = textbox.get("group_name")
                        group_index = textbox.get("group_index")
                        child_path = textbox.get("child_path")

                        if not child_path:
                            child_path = str(textbox.get("child_index", ""))

                        group = None

                        if group_index is not None and 0 <= group_index < len(all_shapes):
                            try:
                                group = all_shapes[group_index]
                            except:
                                pass

                        if not group and group_name:
                            for shape in all_shapes:
                                if shape.name == group_name:
                                    group = shape
                                    break

                        if group and child_path and hasattr(group.api, 'GroupItems'):
                            def navigate_to_child(parent_group, path):
                                path_parts = path.split('/')
                                current_item = parent_group

                                for part in path_parts:
                                    try:
                                        idx = int(part)
                                        if hasattr(current_item.api, 'GroupItems'):
                                            items = current_item.api.GroupItems
                                            if 1 <= idx <= items.Count:
                                                current_item = items.Item(idx)
                                            else:
                                                return None
                                        else:
                                            return None
                                    except:
                                        return None

                                return current_item

                            child_item = navigate_to_child(group, child_path)

                            if child_item:
                                updated = False

                                try:
                                    if hasattr(child_item, 'TextFrame') and child_item.TextFrame.HasText:
                                        child_item.TextFrame.Characters().Text = textbox["value"]
                                        updated = True
                                        textbox_success += 1
                                        app_logger.debug(f"Updated group '{group_name}' child with path {child_path} using TextFrame, count_src {count_src}")
                                except:
                                    pass

                                if not updated:
                                    try:
                                        if hasattr(child_item, 'TextFrame2'):
                                            child_item.TextFrame2.TextRange.Text = textbox["value"]
                                            updated = True
                                            textbox_success += 1
                                            app_logger.debug(f"Updated group '{group_name}' child with path {child_path} using TextFrame2, count_src {count_src}")
                                    except:
                                        pass

                                if not updated:
                                    textbox_failed += 1
                                    app_logger.warning(f"Could not update group '{group_name}' child with path {child_path}, count_src {count_src}")
                            else:
                                textbox_failed += 1
                                app_logger.warning(f"Could not navigate to child with path {child_path} in group '{group_name}', count_src {count_src}")
                        else:
                            textbox_failed += 1
                            app_logger.warning(f"Could not find group '{group_name}' or it lacks GroupItems, count_src {count_src}")
                    except Exception as e:
                        textbox_failed += 1
                        count_src = textbox.get("count_src", "unknown")
                        app_logger.warning(f"Error processing group shape, group: {textbox.get('group_name')}, path: {textbox.get('child_path')}, count_src {count_src}: {str(e)}")
                        continue

                app_logger.info(f"Textbox processing completed for sheet '{sheet_name}': {textbox_success} successful, {textbox_failed} failed")

            except Exception as e:
                app_logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                continue

        # Save the workbook
        result_folder = os.path.join(result_dir)
        os.makedirs(result_folder, exist_ok=True)

        # Use source_lang2target_lang format if available, otherwise fallback to _translated
        if src_lang and dst_lang:
            lang_suffix = f"{src_lang}2{dst_lang}"
        else:
            lang_suffix = "translated"

        result_path = os.path.join(
            result_folder,
            f"{os.path.splitext(os.path.basename(file_path))[0]}_{lang_suffix}{os.path.splitext(file_path)[1]}"
        )

        try:
            wb.save(result_path)
            app_logger.info(f"Translated Excel (without drawing/SmartArt) saved to: {result_path}")
        except Exception as e:
            app_logger.error(f"Failed to save translated Excel: {str(e)}")
            fallback_path = os.path.join(
                result_folder,
                f"{os.path.splitext(os.path.basename(file_path))[0]}_{lang_suffix}_fallback{os.path.splitext(file_path)[1]}"
            )
            try:
                wb.save(fallback_path)
                app_logger.info(f"Translated Excel saved to fallback path: {fallback_path}")
                result_path = fallback_path
            except Exception as e2:
                app_logger.error(f"Failed to save translated Excel to fallback path: {str(e2)}")
                raise

    finally:
        try:
            wb.close()
            app.quit()
        except Exception as e:
            app_logger.error(f"Error closing workbook or quitting app: {str(e)}")

    # Process SmartArt
    if smartart_items:
        app_logger.info(f"Processing {len(smartart_items)} SmartArt translations")
        try:
            if bilingual_mode:
                result_path = _apply_excel_smartart_bilingual_translations_to_file(result_path, smartart_items, translations)
            else:
                result_path = _apply_excel_smartart_translations_to_file(result_path, smartart_items, translations)
        except Exception as e:
            app_logger.error(f"Failed to apply SmartArt translations: {str(e)}")

    # Process Drawing
    if drawing_items:
        app_logger.info(f"Processing {len(drawing_items)} drawing translations")
        try:
            if bilingual_mode:
                result_path = _apply_excel_drawing_bilingual_translations_to_file(result_path, drawing_items, translations)
            else:
                result_path = _apply_excel_drawing_translations_to_file(result_path, drawing_items, translations)
        except Exception as e:
            app_logger.error(f"Failed to apply drawing translations: {str(e)}")

    return result_path


# ============================================================================
# XLWINGS MODE - WRITING HELPERS (Drawing)
# ============================================================================

def _apply_excel_drawing_translations_to_file(file_path: str, drawing_items: List[Dict], translations: Dict) -> str:
    """Apply translations to Excel drawing textboxes."""
    if not drawing_items:
        return file_path

    app_logger.info(f"Processing {len(drawing_items)} Excel drawing translations")

    namespaces = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    }

    items_by_drawing = {}
    for item in drawing_items:
        drawing_index = item['drawing_index']
        if drawing_index not in items_by_drawing:
            items_by_drawing[drawing_index] = []
        items_by_drawing[drawing_index].append(item)

    temp_excel_path = file_path + ".tmp"

    try:
        with ZipFile(file_path, 'r') as original_zip:
            with ZipFile(temp_excel_path, 'w') as new_zip:
                modified_files = set()

                for drawing_index, items in items_by_drawing.items():
                    drawing_path = f"xl/drawings/drawing{drawing_index}.xml"
                    modified_files.add(drawing_path)

                for item in original_zip.infolist():
                    if item.filename not in modified_files:
                        try:
                            new_zip.writestr(item, original_zip.read(item.filename))
                        except Exception as e:
                            app_logger.warning(f"Failed to copy file {item.filename}: {e}")

                for drawing_index, items in items_by_drawing.items():
                    drawing_path = f"xl/drawings/drawing{drawing_index}.xml"

                    if drawing_path in original_zip.namelist():
                        try:
                            drawing_xml = original_zip.read(drawing_path)
                            drawing_tree = etree.fromstring(drawing_xml)

                            for item in items:
                                count = str(item['count_src'])
                                translated_text = translations.get(count)

                                if not translated_text:
                                    app_logger.warning(f"Missing translation for Excel drawing count {count}")
                                    continue

                                translated_text = translated_text.replace("␊", "\n").replace("␍", "\r")

                                try:
                                    xpath = item.get('xpath')
                                    if not xpath:
                                        anchor_idx = item['anchor_index']
                                        tb_idx = item['textbox_index']
                                        tx_body_idx = item['tx_body_index']
                                        p_idx = item['paragraph_index']
                                        xpath = f".//xdr:twoCellAnchor[{anchor_idx + 1}]//xdr:sp[{tb_idx + 1}]//xdr:txBody[{tx_body_idx + 1}]//a:p[{p_idx + 1}]"

                                    paragraphs = drawing_tree.xpath(xpath, namespaces=namespaces)

                                    if paragraphs:
                                        paragraph = paragraphs[0]
                                        _distribute_drawing_text_to_runs(paragraph, translated_text, item, namespaces)
                                        app_logger.info(f"Updated Excel drawing text for drawing {drawing_index}, count_src {count}")
                                    else:
                                        success = _fallback_drawing_text_update(drawing_tree, item, translated_text, namespaces)
                                        if success:
                                            app_logger.info(f"Updated Excel drawing text (fallback) for drawing {drawing_index}, count_src {count}")
                                        else:
                                            app_logger.warning(f"Could not find element to update for drawing {drawing_index}, count_src {count}")

                                except Exception as e:
                                    app_logger.error(f"Failed to update drawing text for count_src {count}: {e}")

                            modified_drawing_xml = etree.tostring(drawing_tree, xml_declaration=True,
                                                                 encoding="UTF-8", standalone="yes")
                            new_zip.writestr(drawing_path, modified_drawing_xml)
                            app_logger.info(f"Saved modified Excel drawing file: {drawing_path}")

                        except Exception as e:
                            app_logger.error(f"Failed to apply Excel drawing translation to {drawing_path}: {e}")
                            try:
                                new_zip.writestr(drawing_path, original_zip.read(drawing_path))
                            except Exception as fallback_e:
                                app_logger.error(f"Failed to copy original drawing file as fallback: {fallback_e}")

        shutil.move(temp_excel_path, file_path)
        app_logger.info(f"Excel drawing translations applied successfully")
        return file_path

    except Exception as e:
        app_logger.error(f"Failed to apply Excel drawing translations: {e}")
        if os.path.exists(temp_excel_path):
            try:
                os.remove(temp_excel_path)
            except:
                pass
        return file_path


def _fallback_drawing_text_update(drawing_tree, item: Dict, translated_text: str, namespaces: Dict) -> bool:
    """Fallback method to find and update drawing text by matching content."""
    try:
        original_text = item.get('original_text', '').strip()
        if not original_text:
            return False

        all_text_nodes = drawing_tree.xpath('.//a:t', namespaces=namespaces)

        for text_node in all_text_nodes:
            if text_node.text and text_node.text.strip() == original_text:
                text_node.text = translated_text
                app_logger.debug(f"Updated text via fallback method: '{original_text}' -> '{translated_text[:50]}...'")
                return True

        for text_node in all_text_nodes:
            if text_node.text and original_text in text_node.text:
                text_node.text = text_node.text.replace(original_text, translated_text)
                app_logger.debug(f"Updated text via partial matching: '{original_text}' -> '{translated_text[:50]}...'")
                return True

        return False

    except Exception as e:
        app_logger.warning(f"Error in fallback text update: {e}")
        return False


def _distribute_drawing_text_to_runs(parent_element, translated_text: str, item: Dict, namespaces: Dict):
    """Distribute translated text across multiple runs in Excel drawing."""
    text_runs = parent_element.xpath('.//a:r', namespaces=namespaces)

    if not text_runs:
        text_nodes = parent_element.xpath('.//a:t', namespaces=namespaces)
        if text_nodes:
            text_nodes[0].text = translated_text
        return

    original_run_texts = item.get('run_texts', [])
    original_run_lengths = item.get('run_lengths', [])

    if not original_run_texts or len(original_run_texts) != len(text_runs):
        app_logger.warning(f"Mismatch in Excel drawing run structure, using simple distribution")
        _simple_drawing_text_distribution(text_runs, translated_text, namespaces)
        return

    _intelligent_drawing_text_distribution(text_runs, translated_text, original_run_texts, original_run_lengths, namespaces)


def _simple_drawing_text_distribution(text_runs, translated_text: str, namespaces: Dict):
    """Simple fallback distribution method for Excel drawing."""
    if not text_runs:
        return

    for i, text_run in enumerate(text_runs):
        text_node = text_run.xpath('./a:t', namespaces=namespaces)
        if text_node:
            if i == 0:
                text_node[0].text = translated_text
            else:
                text_node[0].text = ""


def _intelligent_drawing_text_distribution(text_runs, translated_text: str, original_run_texts: List[str],
                                         original_run_lengths: List[int], namespaces: Dict):
    """Intelligent text distribution for Excel drawing."""
    meaningful_runs = [(i, length) for i, length in enumerate(original_run_lengths) if length > 0]
    total_meaningful_length = sum(length for _, length in meaningful_runs)

    if total_meaningful_length == 0:
        _simple_drawing_text_distribution(text_runs, translated_text, namespaces)
        return

    translated_chars = list(translated_text)
    char_index = 0

    for run_index, text_run in enumerate(text_runs):
        text_node = text_run.xpath('./a:t', namespaces=namespaces)
        if not text_node:
            continue

        original_text = original_run_texts[run_index] if run_index < len(original_run_texts) else ""
        original_length = original_run_lengths[run_index] if run_index < len(original_run_lengths) else 0

        if original_length == 0 or not original_text.strip():
            if original_text and not original_text.strip():
                if char_index < len(translated_chars) and translated_chars[char_index] == ' ':
                    text_node[0].text = ' '
                    char_index += 1
                else:
                    text_node[0].text = ""
            else:
                text_node[0].text = ""
            continue

        if run_index == len(text_runs) - 1:
            remaining_text = ''.join(translated_chars[char_index:])
            text_node[0].text = remaining_text
        else:
            proportion = original_length / total_meaningful_length
            target_length = max(1, int(len(translated_text) * proportion))

            run_text = ""
            chars_taken = 0

            while chars_taken < target_length and char_index < len(translated_chars):
                char = translated_chars[char_index]
                run_text += char
                chars_taken += 1
                char_index += 1

                if chars_taken >= target_length and char_index < len(translated_chars):
                    if char != ' ' and translated_chars[char_index] != ' ':
                        while (char_index < len(translated_chars) and
                               translated_chars[char_index] != ' ' and
                               chars_taken < target_length * 1.5):
                            char = translated_chars[char_index]
                            run_text += char
                            chars_taken += 1
                            char_index += 1
                    break

            text_node[0].text = run_text


# ============================================================================
# XLWINGS MODE - WRITING HELPERS (SmartArt)
# ============================================================================

def _apply_excel_smartart_translations_to_file(file_path: str, smartart_items: List[Dict], translations: Dict) -> str:
    """Apply translations to Excel SmartArt diagrams."""
    if not smartart_items:
        return file_path

    app_logger.info(f"Processing {len(smartart_items)} Excel SmartArt translations")

    namespaces = {
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'dgm': 'http://schemas.openxmlformats.org/drawingml/2006/diagram',
        'dsp': 'http://schemas.microsoft.com/office/drawing/2008/diagram',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    }

    items_by_diagram = {}
    for item in smartart_items:
        diagram_index = item['diagram_index']
        if diagram_index not in items_by_diagram:
            items_by_diagram[diagram_index] = []
        items_by_diagram[diagram_index].append(item)

    temp_excel_path = file_path + ".tmp"

    try:
        with ZipFile(file_path, 'r') as original_zip:
            with ZipFile(temp_excel_path, 'w') as new_zip:
                modified_files = set()

                for diagram_index, items in items_by_diagram.items():
                    drawing_path = f"xl/diagrams/drawing{diagram_index}.xml"
                    data_path = f"xl/diagrams/data{diagram_index}.xml"
                    modified_files.add(drawing_path)
                    modified_files.add(data_path)

                for item in original_zip.infolist():
                    if item.filename not in modified_files:
                        try:
                            new_zip.writestr(item, original_zip.read(item.filename))
                        except Exception as e:
                            app_logger.warning(f"Failed to copy file {item.filename}: {e}")

                for diagram_index, items in items_by_diagram.items():
                    drawing_path = f"xl/diagrams/drawing{diagram_index}.xml"
                    data_path = f"xl/diagrams/data{diagram_index}.xml"

                    if drawing_path in original_zip.namelist():
                        try:
                            drawing_xml = original_zip.read(drawing_path)
                            drawing_tree = etree.fromstring(drawing_xml)

                            for item in items:
                                count = str(item['count_src'])
                                translated_text = translations.get(count)

                                if not translated_text:
                                    app_logger.warning(f"Missing translation for Excel SmartArt count {count}")
                                    continue

                                translated_text = translated_text.replace("␊", "\n").replace("␍", "\r")

                                shapes_with_txbody = drawing_tree.xpath('.//dsp:sp[.//dsp:txBody]', namespaces=namespaces)

                                if item['shape_index'] < len(shapes_with_txbody):
                                    shape = shapes_with_txbody[item['shape_index']]

                                    tx_bodies = shape.xpath('.//dsp:txBody', namespaces=namespaces)
                                    if item['tx_body_index'] < len(tx_bodies):
                                        tx_body = tx_bodies[item['tx_body_index']]

                                        paragraphs = tx_body.xpath('.//a:p', namespaces=namespaces)
                                        if item['paragraph_index'] < len(paragraphs):
                                            paragraph = paragraphs[item['paragraph_index']]
                                            _distribute_excel_smartart_text_to_runs(paragraph, translated_text, item, namespaces)
                                            app_logger.info(f"Updated Excel SmartArt drawing text for diagram {diagram_index}, shape {item['shape_index']}, count_src {count}")

                            modified_drawing_xml = etree.tostring(drawing_tree, xml_declaration=True,
                                                                 encoding="UTF-8", standalone="yes")
                            new_zip.writestr(drawing_path, modified_drawing_xml)
                            app_logger.info(f"Saved modified Excel SmartArt drawing file: {drawing_path}")

                        except Exception as e:
                            app_logger.error(f"Failed to apply Excel SmartArt translation to {drawing_path}: {e}")
                            try:
                                new_zip.writestr(drawing_path, original_zip.read(drawing_path))
                            except Exception as fallback_e:
                                app_logger.error(f"Failed to copy original drawing file as fallback: {fallback_e}")

                    if data_path in original_zip.namelist():
                        try:
                            data_xml = original_zip.read(data_path)
                            data_tree = etree.fromstring(data_xml)

                            for item in items:
                                count = str(item['count_src'])
                                translated_text = translations.get(count)

                                if not translated_text:
                                    continue

                                translated_text = translated_text.replace("␊", "\n").replace("␍", "\r")
                                original_text = item.get('original_text', '')

                                points = data_tree.xpath('.//dgm:pt[.//a:t]', namespaces=namespaces)

                                for point in points:
                                    point_paragraphs = point.xpath('.//a:p', namespaces=namespaces)
                                    for p_idx, point_paragraph in enumerate(point_paragraphs):
                                        point_text_runs = point_paragraph.xpath('.//a:r', namespaces=namespaces)
                                        if point_text_runs:
                                            point_run_info = _process_excel_smartart_text_runs(point_text_runs, namespaces)
                                            if point_run_info['merged_text'].strip() == original_text.strip():
                                                _distribute_excel_smartart_text_to_runs(point_paragraph, translated_text, item, namespaces)
                                                app_logger.info(f"Updated Excel SmartArt data text for diagram {diagram_index}, count_src {count}: '{original_text}' -> '{translated_text[:50]}...'")
                                                break

                            modified_data_xml = etree.tostring(data_tree, xml_declaration=True,
                                                              encoding="UTF-8", standalone="yes")
                            new_zip.writestr(data_path, modified_data_xml)
                            app_logger.info(f"Saved modified Excel SmartArt data file: {data_path}")

                        except Exception as e:
                            app_logger.error(f"Failed to apply Excel SmartArt translation to {data_path}: {e}")
                            try:
                                new_zip.writestr(data_path, original_zip.read(data_path))
                            except Exception as fallback_e:
                                app_logger.error(f"Failed to copy original data file as fallback: {fallback_e}")

        shutil.move(temp_excel_path, file_path)
        app_logger.info(f"Excel SmartArt translations applied successfully")
        return file_path

    except Exception as e:
        app_logger.error(f"Failed to apply Excel SmartArt translations: {e}")
        if os.path.exists(temp_excel_path):
            try:
                os.remove(temp_excel_path)
            except:
                pass
        return file_path


def _distribute_excel_smartart_text_to_runs(parent_element, translated_text: str, item: Dict, namespaces: Dict):
    """Distribute translated text across multiple runs in Excel SmartArt."""
    text_runs = parent_element.xpath('.//a:r', namespaces=namespaces)

    if not text_runs:
        return

    original_run_texts = item.get('run_texts', [])
    original_run_lengths = item.get('run_lengths', [])

    if not original_run_texts or len(original_run_texts) != len(text_runs):
        app_logger.warning(f"Mismatch in Excel SmartArt run structure, using simple distribution")
        _simple_excel_smartart_text_distribution(text_runs, translated_text, namespaces)
        return

    _intelligent_excel_smartart_text_distribution(text_runs, translated_text, original_run_texts, original_run_lengths, namespaces)


def _simple_excel_smartart_text_distribution(text_runs, translated_text: str, namespaces: Dict):
    """Simple fallback distribution method for Excel SmartArt."""
    if not text_runs:
        return

    for i, text_run in enumerate(text_runs):
        text_node = text_run.xpath('./a:t', namespaces=namespaces)
        if text_node:
            if i == 0:
                text_node[0].text = translated_text
            else:
                text_node[0].text = ""


def _intelligent_excel_smartart_text_distribution(text_runs, translated_text: str, original_run_texts: List[str],
                                                 original_run_lengths: List[int], namespaces: Dict):
    """Intelligent text distribution for Excel SmartArt."""
    meaningful_runs = [(i, length) for i, length in enumerate(original_run_lengths) if length > 0]
    total_meaningful_length = sum(length for _, length in meaningful_runs)

    if total_meaningful_length == 0:
        _simple_excel_smartart_text_distribution(text_runs, translated_text, namespaces)
        return

    translated_chars = list(translated_text)
    char_index = 0

    for run_index, text_run in enumerate(text_runs):
        text_node = text_run.xpath('./a:t', namespaces=namespaces)
        if not text_node:
            continue

        original_text = original_run_texts[run_index] if run_index < len(original_run_texts) else ""
        original_length = original_run_lengths[run_index] if run_index < len(original_run_lengths) else 0

        if original_length == 0 or not original_text.strip():
            if original_text and not original_text.strip():
                if char_index < len(translated_chars) and translated_chars[char_index] == ' ':
                    text_node[0].text = ' '
                    char_index += 1
                else:
                    text_node[0].text = ""
            else:
                text_node[0].text = ""
            continue

        if run_index == len(text_runs) - 1:
            remaining_text = ''.join(translated_chars[char_index:])
            text_node[0].text = remaining_text
        else:
            proportion = original_length / total_meaningful_length
            target_length = max(1, int(len(translated_text) * proportion))

            run_text = ""
            chars_taken = 0

            while chars_taken < target_length and char_index < len(translated_chars):
                char = translated_chars[char_index]
                run_text += char
                chars_taken += 1
                char_index += 1

                if chars_taken >= target_length and char_index < len(translated_chars):
                    if char != ' ' and translated_chars[char_index] != ' ':
                        while (char_index < len(translated_chars) and
                               translated_chars[char_index] != ' ' and
                               chars_taken < target_length * 1.5):
                            char = translated_chars[char_index]
                            run_text += char
                            chars_taken += 1
                            char_index += 1
                    break

            text_node[0].text = run_text


# ============================================================================
# BILINGUAL MODE - WRITING HELPERS (Drawing)
# ============================================================================

def _apply_excel_drawing_bilingual_translations_to_file(file_path: str, drawing_items: List[Dict], translations: Dict) -> str:
    """Apply bilingual translations to Excel drawing textboxes."""
    if not drawing_items:
        return file_path

    app_logger.info(f"Processing {len(drawing_items)} Excel drawing translations with bilingual format")

    namespaces = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    }

    items_by_drawing = {}
    for item in drawing_items:
        drawing_index = item['drawing_index']
        if drawing_index not in items_by_drawing:
            items_by_drawing[drawing_index] = []
        items_by_drawing[drawing_index].append(item)

    temp_excel_path = file_path + ".tmp"

    try:
        with ZipFile(file_path, 'r') as original_zip:
            with ZipFile(temp_excel_path, 'w') as new_zip:
                modified_files = set()

                for drawing_index, items in items_by_drawing.items():
                    drawing_path = f"xl/drawings/drawing{drawing_index}.xml"
                    modified_files.add(drawing_path)

                for item in original_zip.infolist():
                    if item.filename not in modified_files:
                        try:
                            new_zip.writestr(item, original_zip.read(item.filename))
                        except Exception as e:
                            app_logger.warning(f"Failed to copy file {item.filename}: {e}")

                for drawing_index, items in items_by_drawing.items():
                    drawing_path = f"xl/drawings/drawing{drawing_index}.xml"

                    if drawing_path in original_zip.namelist():
                        try:
                            drawing_xml = original_zip.read(drawing_path)
                            drawing_tree = etree.fromstring(drawing_xml)

                            for item in items:
                                count = str(item['count_src'])
                                translated_text = translations.get(count)

                                if not translated_text:
                                    app_logger.warning(f"Missing translation for Excel drawing count {count}")
                                    continue

                                # Format as bilingual text
                                bilingual_text = _format_bilingual_text(
                                    item['value'], translated_text, "drawing"
                                )

                                try:
                                    xpath = item.get('xpath')
                                    if not xpath:
                                        anchor_idx = item['anchor_index']
                                        tb_idx = item['textbox_index']
                                        tx_body_idx = item['tx_body_index']
                                        p_idx = item['paragraph_index']
                                        xpath = f".//xdr:twoCellAnchor[{anchor_idx + 1}]//xdr:sp[{tb_idx + 1}]//xdr:txBody[{tx_body_idx + 1}]//a:p[{p_idx + 1}]"

                                    paragraphs = drawing_tree.xpath(xpath, namespaces=namespaces)

                                    if paragraphs:
                                        paragraph = paragraphs[0]
                                        _simple_drawing_text_distribution(paragraph.xpath('.//a:r', namespaces=namespaces), bilingual_text, namespaces)
                                        app_logger.info(f"Updated bilingual Excel drawing text for drawing {drawing_index}, count_src {count}")
                                    else:
                                        success = _fallback_drawing_text_update(drawing_tree, item, bilingual_text, namespaces)
                                        if success:
                                            app_logger.info(f"Updated bilingual Excel drawing text (fallback) for drawing {drawing_index}, count_src {count}")
                                        else:
                                            app_logger.warning(f"Could not find element to update for bilingual drawing {drawing_index}, count_src {count}")

                                except Exception as e:
                                    app_logger.error(f"Failed to update bilingual drawing text for count_src {count}: {e}")

                            modified_drawing_xml = etree.tostring(drawing_tree, xml_declaration=True,
                                                                 encoding="UTF-8", standalone="yes")
                            new_zip.writestr(drawing_path, modified_drawing_xml)
                            app_logger.info(f"Saved modified bilingual Excel drawing file: {drawing_path}")

                        except Exception as e:
                            app_logger.error(f"Failed to apply bilingual Excel drawing translation to {drawing_path}: {e}")
                            try:
                                new_zip.writestr(drawing_path, original_zip.read(drawing_path))
                            except Exception as fallback_e:
                                app_logger.error(f"Failed to copy original drawing file as fallback: {fallback_e}")

        shutil.move(temp_excel_path, file_path)
        app_logger.info(f"Excel bilingual drawing translations applied successfully")
        return file_path

    except Exception as e:
        app_logger.error(f"Failed to apply bilingual Excel drawing translations: {e}")
        if os.path.exists(temp_excel_path):
            try:
                os.remove(temp_excel_path)
            except:
                pass
        return file_path


# ============================================================================
# BILINGUAL MODE - WRITING HELPERS (SmartArt)
# ============================================================================

def _apply_excel_smartart_bilingual_translations_to_file(file_path: str, smartart_items: List[Dict], translations: Dict) -> str:
    """Apply bilingual translations to Excel SmartArt diagrams."""
    if not smartart_items:
        return file_path

    app_logger.info(f"Processing {len(smartart_items)} Excel SmartArt translations with bilingual format")

    namespaces = {
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'dgm': 'http://schemas.openxmlformats.org/drawingml/2006/diagram',
        'dsp': 'http://schemas.microsoft.com/office/drawing/2008/diagram',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    }

    items_by_diagram = {}
    for item in smartart_items:
        diagram_index = item['diagram_index']
        if diagram_index not in items_by_diagram:
            items_by_diagram[diagram_index] = []
        items_by_diagram[diagram_index].append(item)

    temp_excel_path = file_path + ".tmp"

    try:
        with ZipFile(file_path, 'r') as original_zip:
            with ZipFile(temp_excel_path, 'w') as new_zip:
                modified_files = set()

                for diagram_index, items in items_by_diagram.items():
                    drawing_path = f"xl/diagrams/drawing{diagram_index}.xml"
                    data_path = f"xl/diagrams/data{diagram_index}.xml"
                    modified_files.add(drawing_path)
                    modified_files.add(data_path)

                for item in original_zip.infolist():
                    if item.filename not in modified_files:
                        try:
                            new_zip.writestr(item, original_zip.read(item.filename))
                        except Exception as e:
                            app_logger.warning(f"Failed to copy file {item.filename}: {e}")

                for diagram_index, items in items_by_diagram.items():
                    drawing_path = f"xl/diagrams/drawing{diagram_index}.xml"
                    data_path = f"xl/diagrams/data{diagram_index}.xml"

                    if drawing_path in original_zip.namelist():
                        try:
                            drawing_xml = original_zip.read(drawing_path)
                            drawing_tree = etree.fromstring(drawing_xml)

                            for item in items:
                                count = str(item['count_src'])
                                translated_text = translations.get(count)

                                if not translated_text:
                                    app_logger.warning(f"Missing translation for Excel SmartArt count {count}")
                                    continue

                                # Format as bilingual text
                                bilingual_text = _format_bilingual_text(
                                    item['value'], translated_text, "smartart"
                                )

                                shapes_with_txbody = drawing_tree.xpath('.//dsp:sp[.//dsp:txBody]', namespaces=namespaces)

                                if item['shape_index'] < len(shapes_with_txbody):
                                    shape = shapes_with_txbody[item['shape_index']]

                                    tx_bodies = shape.xpath('.//dsp:txBody', namespaces=namespaces)
                                    if item['tx_body_index'] < len(tx_bodies):
                                        tx_body = tx_bodies[item['tx_body_index']]

                                        paragraphs = tx_body.xpath('.//a:p', namespaces=namespaces)
                                        if item['paragraph_index'] < len(paragraphs):
                                            paragraph = paragraphs[item['paragraph_index']]
                                            _simple_excel_smartart_text_distribution(paragraph.xpath('.//a:r', namespaces=namespaces), bilingual_text, namespaces)
                                            app_logger.info(f"Updated bilingual Excel SmartArt drawing text for diagram {diagram_index}, shape {item['shape_index']}, count_src {count}")

                            modified_drawing_xml = etree.tostring(drawing_tree, xml_declaration=True,
                                                                 encoding="UTF-8", standalone="yes")
                            new_zip.writestr(drawing_path, modified_drawing_xml)
                            app_logger.info(f"Saved modified bilingual Excel SmartArt drawing file: {drawing_path}")

                        except Exception as e:
                            app_logger.error(f"Failed to apply bilingual Excel SmartArt translation to {drawing_path}: {e}")
                            try:
                                new_zip.writestr(drawing_path, original_zip.read(drawing_path))
                            except Exception as fallback_e:
                                app_logger.error(f"Failed to copy original drawing file as fallback: {fallback_e}")

                    if data_path in original_zip.namelist():
                        try:
                            data_xml = original_zip.read(data_path)
                            data_tree = etree.fromstring(data_xml)

                            for item in items:
                                count = str(item['count_src'])
                                translated_text = translations.get(count)

                                if not translated_text:
                                    continue

                                # Format as bilingual text
                                bilingual_text = _format_bilingual_text(
                                    item['value'], translated_text, "smartart"
                                )
                                original_text = item.get('original_text', '')

                                points = data_tree.xpath('.//dgm:pt[.//a:t]', namespaces=namespaces)

                                for point in points:
                                    point_paragraphs = point.xpath('.//a:p', namespaces=namespaces)
                                    for p_idx, point_paragraph in enumerate(point_paragraphs):
                                        point_text_runs = point_paragraph.xpath('.//a:r', namespaces=namespaces)
                                        if point_text_runs:
                                            point_run_info = _process_excel_smartart_text_runs(point_text_runs, namespaces)
                                            if point_run_info['merged_text'].strip() == original_text.strip():
                                                _simple_excel_smartart_text_distribution(point_text_runs, bilingual_text, namespaces)
                                                app_logger.info(f"Updated bilingual Excel SmartArt data text for diagram {diagram_index}, count_src {count}: '{original_text}' -> bilingual format")
                                                break

                            modified_data_xml = etree.tostring(data_tree, xml_declaration=True,
                                                              encoding="UTF-8", standalone="yes")
                            new_zip.writestr(data_path, modified_data_xml)
                            app_logger.info(f"Saved modified bilingual Excel SmartArt data file: {data_path}")

                        except Exception as e:
                            app_logger.error(f"Failed to apply bilingual Excel SmartArt translation to {data_path}: {e}")
                            try:
                                new_zip.writestr(data_path, original_zip.read(data_path))
                            except Exception as fallback_e:
                                app_logger.error(f"Failed to copy original data file as fallback: {fallback_e}")

        shutil.move(temp_excel_path, file_path)
        app_logger.info(f"Excel bilingual SmartArt translations applied successfully")
        return file_path

    except Exception as e:
        app_logger.error(f"Failed to apply bilingual Excel SmartArt translations: {e}")
        if os.path.exists(temp_excel_path):
            try:
                os.remove(temp_excel_path)
            except:
                pass
        return file_path
