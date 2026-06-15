# Corpus tests: XLSX structures.
#   - multiple merged regions        - cross-sheet formula
#   - date/percent/currency cells untouched, number formats preserved
#   - cell comments                  - frozen panes
#   - CJK sheet names                - long multi-line cell
#
# Run from the repo root:
#   python tests/test_corpus_xlsx.py
import datetime
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from tests.corpus_common import T, check, fake_translate, run, work_dirs

WORK_DIR, TEMP_DIR, RESULT_DIR = work_dirs("xlsx")

LONG_NOTE = ("First line of an unusually long cell note that goes on for a while\n"
             "second line continues the explanation with more detail\n"
             "third line wraps everything up after a considerable amount of text")


def test_xlsx_structures():
    print("XLSX: merges, cross-sheet formula, typed cells, comments, panes, CJK sheets")
    import openpyxl
    from openpyxl.comments import Comment
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)

    src = os.path.join(WORK_DIR, "structures.xlsx")
    wb = openpyxl.Workbook()

    # Sheet 1: "DATA2024" is a code-like name the skip filter leaves alone,
    # so the cross-sheet formula that references it stays valid after the
    # translated workbook is saved.
    ws1 = wb.active
    ws1.title = "DATA2024"
    ws1["A1"] = "Wide merged banner title"
    ws1.merge_cells("A1:D1")
    ws1["A3"] = "Tall merged side label"
    ws1.merge_cells("A3:B5")
    ws1["E1"] = "Second merged region"
    ws1.merge_cells("E1:E3")
    ws1["B7"] = 1999.5                                   # plain number
    ws1["B8"] = datetime.datetime(2024, 3, 15)           # date
    ws1["B8"].number_format = "yyyy-mm-dd"
    ws1["B9"] = 0.375                                    # percent
    ws1["B9"].number_format = "0.00%"
    ws1["B10"] = 1234.5                                  # currency
    ws1["B10"].number_format = '"$"#,##0.00'
    ws1["A12"] = LONG_NOTE                               # long multi-line cell
    ws1["C2"] = "Cell carrying a comment"
    ws1["C2"].comment = Comment("Reviewer remark stays as a comment", "Reviewer")
    ws1.freeze_panes = "B2"

    # Sheet 2: CJK sheet name (gets translated/renamed) + cross-sheet formula
    ws2 = wb.create_sheet("売上データ")
    ws2["A1"] = "Cross sheet total label"
    ws2["B1"] = "=DATA2024!B7*2"
    ws2.freeze_panes = "A2"

    wb.save(src)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="ja", dst_lang="en")

    wb2 = openpyxl.load_workbook(out)
    sheets = wb2.sheetnames
    s1 = wb2["DATA2024"]

    # --- sheet names ---
    check("code-like sheet name untouched", "DATA2024" in sheets, str(sheets))
    # Excel forbids []:*?/\ in sheet names, so the fake [T] marker is
    # sanitized to -T- by design; the rename itself is what matters.
    check("CJK sheet name translated (renamed, brackets sanitized)",
          "-T-売上データ" in sheets, str(sheets))
    s2 = wb2["-T-売上データ"] if "-T-売上データ" in sheets else wb2[sheets[1]]

    # --- text content ---
    check("merged banner translated", s1["A1"].value == T + "Wide merged banner title",
          repr(s1["A1"].value))
    check("all three merged regions intact",
          sorted(str(r) for r in s1.merged_cells.ranges) == ["A1:D1", "A3:B5", "E1:E3"],
          str(list(s1.merged_cells.ranges)))
    check("other merged texts translated",
          s1["A3"].value == T + "Tall merged side label"
          and s1["E1"].value == T + "Second merged region",
          f"{s1['A3'].value!r} / {s1['E1'].value!r}")
    check("long multi-line cell translated with newlines intact",
          s1["A12"].value == T + LONG_NOTE, repr(s1["A12"].value))
    check("second sheet label translated",
          s2["A1"].value == T + "Cross sheet total label", repr(s2["A1"].value))

    # --- typed cells untouched, formats preserved ---
    check("plain number untouched", s1["B7"].value == 1999.5, repr(s1["B7"].value))
    check("date cell untouched with format preserved",
          s1["B8"].value == datetime.datetime(2024, 3, 15)
          and s1["B8"].number_format == "yyyy-mm-dd",
          f"{s1['B8'].value!r} / {s1['B8'].number_format!r}")
    check("percent cell untouched with format preserved",
          s1["B9"].value == 0.375 and s1["B9"].number_format == "0.00%",
          f"{s1['B9'].value!r} / {s1['B9'].number_format!r}")
    check("currency cell untouched with format preserved",
          s1["B10"].value == 1234.5 and s1["B10"].number_format == '"$"#,##0.00',
          f"{s1['B10'].value!r} / {s1['B10'].number_format!r}")

    # --- cross-sheet formula ---
    # The referenced sheet name is non-translatable, so the formula is still
    # valid. (Known limitation: if a referenced sheet IS renamed by
    # translation, formulas are not rewritten to the new name.)
    check("cross-sheet formula preserved verbatim",
          s2["B1"].value == "=DATA2024!B7*2", repr(s2["B1"].value))

    # --- comments / panes ---
    # Known limitation: comment text is not extracted for translation; it
    # must simply survive the round trip untouched.
    check("comment survives untouched",
          s1["C2"].comment is not None
          and "Reviewer remark stays as a comment" in s1["C2"].comment.text,
          repr(s1["C2"].comment))
    check("comment-bearing cell text translated",
          s1["C2"].value == T + "Cell carrying a comment", repr(s1["C2"].value))
    check("frozen panes preserved on both sheets",
          s1.freeze_panes == "B2" and s2.freeze_panes == "A2",
          f"{s1.freeze_panes} / {s2.freeze_panes}")


_DRAWING_XML = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
                ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
                '<xdr:twoCellAnchor>'
                '<xdr:from><xdr:col>2</xdr:col><xdr:colOff>0</xdr:colOff>'
                '<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
                '<xdr:to><xdr:col>5</xdr:col><xdr:colOff>0</xdr:colOff>'
                '<xdr:row>6</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
                '<xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="TextBox 1"/>'
                '<xdr:cNvSpPr txBox="1"/></xdr:nvSpPr><xdr:spPr/>'
                '<xdr:txBody><a:bodyPr/><a:p><a:r><a:rPr lang="en-US"/>'
                '<a:t>Standalone textbox caption</a:t></a:r></a:p></xdr:txBody></xdr:sp>'
                '<xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>')


def _inject_textbox(xlsx_path):
    """Add a real DrawingML textbox to an existing .xlsx (openpyxl can't author one)."""
    import shutil
    import zipfile
    tmp = xlsx_path + ".inj"
    with zipfile.ZipFile(xlsx_path) as zin, zipfile.ZipFile(tmp, "w") as zout:
        for n in zin.namelist():
            data = zin.read(n)
            if n == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/drawings/drawing1.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
            if n == "xl/worksheets/sheet1.xml":
                if b"xmlns:r=" not in data.split(b">", 1)[0]:
                    data = data.replace(
                        b"<worksheet ",
                        b'<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        b'officeDocument/2006/relationships" ', 1)
                data = data.replace(b"</worksheet>", b'<drawing r:id="rId1"/></worksheet>')
            zout.writestr(n, data)
        zout.writestr("xl/drawings/drawing1.xml", _DRAWING_XML)
        zout.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
            '2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>')
    shutil.move(tmp, xlsx_path)


def test_xlsx_textbox_openpyxl_path():
    print("XLSX: textbox in a drawing is translated via the default openpyxl path")
    import zipfile
    import openpyxl
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)

    src = os.path.join(WORK_DIR, "textbox.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Plain cell text"
    wb.save(src)
    _inject_textbox(src)

    # use_xlwings=False -> this is the production default path
    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    import json
    with open(src_json, encoding="utf-8") as f:
        extracted = [i["value"] for i in json.load(f)]
    check("textbox caption extracted by the openpyxl path",
          "Standalone textbox caption" in extracted, str(extracted))

    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)

    # cell still translated, file still valid (sheet name itself gets renamed)
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.worksheets[0]
    check("cell text still translated", ws2["A1"].value == T + "Plain cell text",
          repr(ws2["A1"].value))

    # the drawing part survived openpyxl.save and its text was translated
    with zipfile.ZipFile(out) as z:
        names = z.namelist()
        check("drawing part survived openpyxl save",
              "xl/drawings/drawing1.xml" in names, str([n for n in names if "draw" in n]))
        drawing = z.read("xl/drawings/drawing1.xml").decode("utf-8")
    check("textbox caption translated in the drawing XML",
          T + "Standalone textbox caption" in drawing, drawing)


_CHART_XML = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
              '<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
              ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><c:chart>'
              '<c:title><c:tx><c:rich><a:p><a:r><a:t>Quarterly Revenue Title</a:t>'
              '</a:r></a:p></c:rich></c:tx></c:title><c:plotArea><c:barChart><c:ser>'
              '<c:cat><c:strRef><c:f>Sheet1!$A$1</c:f><c:strCache><c:ptCount val="1"/>'
              '<c:pt idx="0"><c:v>Region Category Label</c:v></c:pt></c:strCache>'
              '</c:strRef></c:cat></c:ser></c:barChart></c:plotArea></c:chart></c:chartSpace>')


def _inject_chart(xlsx_path):
    import shutil
    import zipfile
    tmp = xlsx_path + ".inj"
    with zipfile.ZipFile(xlsx_path) as zin, zipfile.ZipFile(tmp, "w") as zout:
        for n in zin.namelist():
            data = zin.read(n)
            if n == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/charts/chart1.xml" ContentType='
                    b'"application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/></Types>')
            zout.writestr(n, data)
        zout.writestr("xl/charts/chart1.xml", _CHART_XML)
    shutil.move(tmp, xlsx_path)


def test_xlsx_chart_openpyxl_path():
    print("XLSX: chart title + category cache translated via the openpyxl path")
    import zipfile
    import openpyxl
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)

    src = os.path.join(WORK_DIR, "chart.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "Region Category Label"
    wb.save(src)
    _inject_chart(src)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    import json
    with open(src_json, encoding="utf-8") as f:
        extracted = [i["value"] for i in json.load(f)]
    check("chart title extracted", "Quarterly Revenue Title" in extracted, str(extracted))
    check("chart category cache extracted", "Region Category Label" in extracted, str(extracted))

    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)
    with zipfile.ZipFile(out) as z:
        check("chart part survived", "xl/charts/chart1.xml" in z.namelist(), str(z.namelist()))
        chart = z.read("xl/charts/chart1.xml").decode("utf-8")
    check("chart title translated", T + "Quarterly Revenue Title" in chart, chart)
    check("chart category cache translated", T + "Region Category Label" in chart, chart)


def test_xlsx_comment_openpyxl_path():
    print("XLSX: cell comment text translated via the openpyxl path")
    import zipfile
    import openpyxl
    from openpyxl.comments import Comment
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)

    src = os.path.join(WORK_DIR, "comment.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Cell body text"
    ws["A1"].comment = Comment("Reviewer remark to translate", "Rev")
    wb.save(src)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    import json
    with open(src_json, encoding="utf-8") as f:
        extracted = [i["value"] for i in json.load(f)]
    check("comment text extracted", any("Reviewer remark to translate" in v for v in extracted),
          str(extracted))

    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)
    with zipfile.ZipFile(out) as z:
        cfiles = [n for n in z.namelist() if "comment" in n.lower() and n.endswith(".xml")]
        check("comment part present", bool(cfiles), str(z.namelist()))
        comments = "".join(z.read(n).decode("utf-8") for n in cfiles)
    check("comment text translated", T + "Reviewer remark to translate" in comments, comments)


def _inject_mixed_drawing(xlsx_path, png_bytes):
    """Add a drawing with BOTH a picture and a textbox (openpyxl rewrites such
    a drawing lossily, dropping the textbox)."""
    import shutil
    import zipfile
    draw = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<xdr:twoCellAnchor><xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>'
            '<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:to><xdr:col>3</xdr:col>'
            '<xdr:colOff>0</xdr:colOff><xdr:row>5</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
            '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="2" name="Picture 1"/><xdr:cNvPicPr/></xdr:nvPicPr>'
            '<xdr:blipFill><a:blip r:embed="rId1"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            '<xdr:spPr/></xdr:pic><xdr:clientData/></xdr:twoCellAnchor>'
            '<xdr:twoCellAnchor><xdr:from><xdr:col>4</xdr:col><xdr:colOff>0</xdr:colOff>'
            '<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:to><xdr:col>7</xdr:col>'
            '<xdr:colOff>0</xdr:colOff><xdr:row>5</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
            '<xdr:sp><xdr:nvSpPr><xdr:cNvPr id="3" name="TextBox 1"/><xdr:cNvSpPr txBox="1"/>'
            '</xdr:nvSpPr><xdr:spPr/><xdr:txBody><a:bodyPr/><a:p><a:r><a:t>Caption beside picture</a:t>'
            '</a:r></a:p></xdr:txBody></xdr:sp><xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>')
    tmp = xlsx_path + ".inj"
    with zipfile.ZipFile(xlsx_path) as zin, zipfile.ZipFile(tmp, "w") as zout:
        for n in zin.namelist():
            data = zin.read(n)
            if n == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Default Extension="png" ContentType="image/png"/>'
                    b'<Override PartName="/xl/drawings/drawing1.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
            if n == "xl/worksheets/sheet1.xml":
                if b"xmlns:r=" not in data.split(b">", 1)[0]:
                    data = data.replace(b"<worksheet ",
                                        b'<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                                        b'officeDocument/2006/relationships" ', 1)
                data = data.replace(b"</worksheet>", b'<drawing r:id="rId1"/></worksheet>')
            zout.writestr(n, data)
        zout.writestr("xl/drawings/drawing1.xml", draw)
        zout.writestr("xl/drawings/_rels/drawing1.xml.rels",
                      '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                      '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                      '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
                      '2006/relationships/image" Target="../media/image1.png"/></Relationships>')
        zout.writestr("xl/media/image1.png", png_bytes)
        zout.writestr("xl/worksheets/_rels/sheet1.xml.rels",
                      '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                      '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                      '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
                      '2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>')
    shutil.move(tmp, xlsx_path)


def test_xlsx_mixed_drawing_image_and_textbox():
    print("XLSX: drawing with picture + textbox keeps the IMAGE and translates the textbox")
    import zipfile
    import openpyxl
    from PIL import Image
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)

    png = os.path.join(WORK_DIR, "pic.png")
    Image.new("RGB", (50, 30), (30, 140, 90)).save(png)
    png_bytes = open(png, "rb").read()

    src = os.path.join(WORK_DIR, "mixed.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "Body cell"
    wb.save(src)
    _inject_mixed_drawing(src, png_bytes)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)

    with zipfile.ZipFile(out) as z:
        names = z.namelist()
        check("embedded image survived (bytes identical)",
              "xl/media/image1.png" in names and z.read("xl/media/image1.png") == png_bytes,
              str([n for n in names if "media" in n]))
        drawing = z.read("xl/drawings/drawing1.xml").decode("utf-8") if "xl/drawings/drawing1.xml" in names else ""
    check("picture kept in the drawing (not dropped by openpyxl rewrite)",
          "r:embed" in drawing or "blip" in drawing, drawing[:200])
    check("textbox in the mixed drawing translated",
          T + "Caption beside picture" in drawing, drawing)


def test_xlsx_cell_value_sanitized():
    print("XLSX: translated cell starting with '=' is not turned into a formula")
    import openpyxl
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)
    import json

    src = os.path.join(WORK_DIR, "sanitize.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Data2024"  # non-translatable sheet name
    wb.active["A1"] = "equals note"
    wb.save(src)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    data = json.load(open(src_json, encoding="utf-8"))
    for it in data:
        it["translated"] = "=DANGER()" if it["value"] == "equals note" else "[T]" + it["value"]
    dj = os.path.join(TEMP_DIR, "sanitize", "dst.json")
    json.dump(data, open(dj, "w", encoding="utf-8"))
    out = write_translated_content_to_excel(src, src_json, dj, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)

    wb2 = openpyxl.load_workbook(out)
    cell = wb2["Data2024"]["A1"]
    check("'='-leading translation stored as literal text, not a formula",
          cell.data_type != "f" and str(cell.value).strip() == "=DANGER()",
          f"data_type={cell.data_type} value={cell.value!r}")


def test_xlsx_data_validation():
    print("XLSX: data-validation prompt/error popups translated, rule intact")
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)

    src = os.path.join(WORK_DIR, "validation.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form"
    ws["A1"] = "Enter a value below"
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="100")
    dv.promptTitle = "Input required"
    dv.prompt = "Please enter a whole number from one to one hundred"
    dv.errorTitle = "Invalid entry"
    dv.error = "The value must be between one and one hundred"
    dv.showInputMessage = True
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add("A2")
    wb.save(src)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)

    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.worksheets[0]  # sheet name "Form" is translatable -> renamed
    dvs = ws2.data_validations.dataValidation
    check("one data validation preserved", len(dvs) == 1, str(len(dvs)))
    d = dvs[0]
    check("promptTitle translated", d.promptTitle == T + "Input required", repr(d.promptTitle))
    check("prompt body translated",
          d.prompt == T + "Please enter a whole number from one to one hundred", repr(d.prompt))
    check("errorTitle translated", d.errorTitle == T + "Invalid entry", repr(d.errorTitle))
    check("error body translated",
          d.error == T + "The value must be between one and one hundred", repr(d.error))
    check("validation rule intact (type + range)",
          d.type == "whole" and str(d.sqref) == "A2", f"{d.type}/{d.sqref}")
    check("cell text still translated", ws2["A1"].value == T + "Enter a value below",
          repr(ws2["A1"].value))


_ALTTEXT_DRAWING_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
    ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
    '<xdr:twoCellAnchor>'
    '<xdr:from><xdr:col>2</xdr:col><xdr:colOff>0</xdr:colOff>'
    '<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
    '<xdr:to><xdr:col>5</xdr:col><xdr:colOff>0</xdr:colOff>'
    '<xdr:row>6</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
    '<xdr:sp><xdr:nvSpPr>'
    '<xdr:cNvPr id="2" name="TextBox 1" title="Diagram alt title"'
    ' descr="A descriptive caption for screen readers"/>'
    '<xdr:cNvSpPr txBox="1"/></xdr:nvSpPr><xdr:spPr/>'
    '<xdr:txBody><a:bodyPr/><a:p><a:r><a:rPr lang="en-US"/>'
    '<a:t>Shape body text</a:t></a:r></a:p></xdr:txBody></xdr:sp>'
    '<xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>')


def _inject_alttext_drawing(xlsx_path):
    """Add a DrawingML shape carrying cNvPr @title/@descr alt-text attributes."""
    import shutil
    import zipfile
    tmp = xlsx_path + ".inj"
    with zipfile.ZipFile(xlsx_path) as zin, zipfile.ZipFile(tmp, "w") as zout:
        for n in zin.namelist():
            data = zin.read(n)
            if n == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/drawings/drawing1.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
            if n == "xl/worksheets/sheet1.xml":
                if b"xmlns:r=" not in data.split(b">", 1)[0]:
                    data = data.replace(
                        b"<worksheet ",
                        b'<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        b'officeDocument/2006/relationships" ', 1)
                data = data.replace(b"</worksheet>", b'<drawing r:id="rId1"/></worksheet>')
            zout.writestr(n, data)
        zout.writestr("xl/drawings/drawing1.xml", _ALTTEXT_DRAWING_XML)
        zout.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
            '2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>')
    shutil.move(tmp, xlsx_path)


def test_xlsx_drawing_alttext():
    print("XLSX: drawing shape alt-text (cNvPr @descr/@title) translated as attributes")
    import json
    import zipfile
    import openpyxl
    from lxml import etree
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)

    src = os.path.join(WORK_DIR, "alttext.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "Plain cell text"
    wb.save(src)
    _inject_alttext_drawing(src)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    with open(src_json, encoding="utf-8") as f:
        data = json.load(f)
    extracted = [i["value"] for i in data]
    check("alt-text title extracted", "Diagram alt title" in extracted, str(extracted))
    check("alt-text descr extracted",
          "A descriptive caption for screen readers" in extracted, str(extracted))
    check("alt-text items typed excel_alttext",
          all(i.get("type") == "excel_alttext"
              for i in data if i["value"] in ("Diagram alt title",
                                              "A descriptive caption for screen readers")),
          str([i.get("type") for i in data]))

    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)

    with zipfile.ZipFile(out) as z:
        names = z.namelist()
        check("drawing part survived", "xl/drawings/drawing1.xml" in names,
              str([n for n in names if "draw" in n]))
        drawing = z.read("xl/drawings/drawing1.xml").decode("utf-8")
    # attributes translated
    check("alt-text title attribute translated",
          T + "Diagram alt title" in drawing, drawing)
    check("alt-text descr attribute translated",
          T + "A descriptive caption for screen readers" in drawing, drawing)
    # still a valid cNvPr structure with both attrs present
    tree = etree.fromstring(drawing.encode("utf-8"))
    ns = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
    cnvpr = tree.xpath(".//xdr:cNvPr", namespaces=ns)[0]
    check("title/descr remain attributes (not element text)",
          cnvpr.get("title") == T + "Diagram alt title"
          and cnvpr.get("descr") == T + "A descriptive caption for screen readers",
          f"{cnvpr.get('title')!r} / {cnvpr.get('descr')!r}")
    check("shape body text also translated", T + "Shape body text" in drawing, drawing)


_THREADED_COMMENT_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<ThreadedComments xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments">'
    '<threadedComment ref="A1" dT="2024-01-01T00:00:00.00" personId="{00000000-0000-0000-0000-000000000001}"'
    ' id="{11111111-1111-1111-1111-111111111111}">'
    '<text>Modern threaded comment to translate</text></threadedComment>'
    '<threadedComment ref="A1" dT="2024-01-01T00:01:00.00" personId="{00000000-0000-0000-0000-000000000001}"'
    ' id="{22222222-2222-2222-2222-222222222222}" parentId="{11111111-1111-1111-1111-111111111111}">'
    '<text>A reply also needing translation</text></threadedComment>'
    '</ThreadedComments>')


def _inject_threaded_comments(xlsx_path):
    """Add a modern threaded-comments part to an existing .xlsx."""
    import shutil
    import zipfile
    tmp = xlsx_path + ".inj"
    with zipfile.ZipFile(xlsx_path) as zin, zipfile.ZipFile(tmp, "w") as zout:
        for n in zin.namelist():
            data = zin.read(n)
            if n == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/threadedComments/threadedComment1.xml" '
                    b'ContentType="application/vnd.ms-excel.threadedcomments+xml"/></Types>')
            zout.writestr(n, data)
        zout.writestr("xl/threadedComments/threadedComment1.xml", _THREADED_COMMENT_XML)
    shutil.move(tmp, xlsx_path)


def test_xlsx_threaded_comment():
    print("XLSX: modern threaded comments translated at the ZIP level")
    import json
    import zipfile
    import openpyxl
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel)

    src = os.path.join(WORK_DIR, "threaded.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "Cell with a threaded comment"
    wb.save(src)
    _inject_threaded_comments(src)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    with open(src_json, encoding="utf-8") as f:
        data = json.load(f)
    extracted = [i["value"] for i in data]
    check("threaded comment extracted",
          "Modern threaded comment to translate" in extracted, str(extracted))
    check("threaded reply extracted",
          "A reply also needing translation" in extracted, str(extracted))
    check("threaded items typed excel_threadedcomment",
          all(i.get("type") == "excel_threadedcomment"
              for i in data if i["value"] in ("Modern threaded comment to translate",
                                              "A reply also needing translation")),
          str([i.get("type") for i in data]))

    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)

    with zipfile.ZipFile(out) as z:
        names = z.namelist()
        check("threaded-comments part survived",
              "xl/threadedComments/threadedComment1.xml" in names,
              str([n for n in names if "hread" in n]))
        tc = z.read("xl/threadedComments/threadedComment1.xml").decode("utf-8")
    check("threaded comment translated",
          T + "Modern threaded comment to translate" in tc, tc)
    check("threaded reply translated",
          T + "A reply also needing translation" in tc, tc)
    # structure survived: reply still references its parent thread
    check("thread structure preserved (parentId intact)",
          'parentId="{11111111-1111-1111-1111-111111111111}"' in tc, tc)


def test_xlsx_header_footer():
    print("XLSX: header/footer literal text translated, &-format codes preserved")
    import zipfile
    import openpyxl
    from core.pipelines.excel_translation_pipeline import (
        extract_excel_content_to_json, write_translated_content_to_excel,
        _split_header_footer)

    src = os.path.join(WORK_DIR, "headerfooter.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Body cell text"
    # &P = page number, &N = total pages, &D = date, &L/&C/&R = position codes.
    ws.oddHeader.center.text = "Confidential &P of &N"
    ws.oddFooter.right.text = "Printed on &D"
    wb.save(src)

    src_json = extract_excel_content_to_json(src, TEMP_DIR, use_xlwings=False)
    import json
    with open(src_json, encoding="utf-8") as f:
        data = json.load(f)
    extracted = [i["value"] for i in data]
    # The literal runs (not the &-codes) are what gets extracted.
    check("header literal 'Confidential ' extracted",
          any(v.strip() == "Confidential" for v in extracted) or "Confidential " in extracted,
          str(extracted))
    check("header literal ' of ' extracted", any(v == " of " for v in extracted),
          str(extracted))
    check("footer literal 'Printed on ' extracted",
          any(v == "Printed on " for v in extracted), str(extracted))
    # The &-codes must NOT be extracted as translatable items.
    check("no &-code leaked into translatable items",
          not any("&" in v for v in extracted), str(extracted))

    dst_json = fake_translate(src_json)
    out = write_translated_content_to_excel(src, src_json, dst_json, RESULT_DIR,
                                            src_lang="en", dst_lang="ja", use_xlwings=False)

    # Re-read header/footer via openpyxl (it round-trips the strings).
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2.worksheets[0]
    header = ws2.oddHeader.center.text
    footer = ws2.oddFooter.right.text
    check("header literal translated", T + "Confidential" in header, repr(header))
    check("header ' of ' literal translated", T + " of " in header, repr(header))
    # &-codes survive intact AND in order
    check("header &P and &N codes survive in order",
          "&P" in header and "&N" in header
          and header.index("&P") < header.index("&N"),
          repr(header))
    check("footer literal translated and &D code survives",
          T + "Printed on " in footer and "&D" in footer, repr(footer))

    # The splitter itself: codes preserved, only literal runs are 'text'.
    segs = _split_header_footer('&C&"Arial,Bold"&12Title &P/&N')
    codes = [v for k, v in segs if k == "code"]
    texts = [v for k, v in segs if k == "text"]
    check("splitter isolates &-codes",
          codes == ['&C', '&"Arial,Bold"', '&12', '&P', '&N'],
          str(codes))
    check("splitter keeps only literal runs as text",
          texts == ['Title ', '/'], str(texts))
    check("splitter is lossless (reassembles original)",
          "".join(v for _, v in segs) == '&C&"Arial,Bold"&12Title &P/&N',
          str(segs))


if __name__ == "__main__":
    run([test_xlsx_structures, test_xlsx_textbox_openpyxl_path,
         test_xlsx_chart_openpyxl_path, test_xlsx_comment_openpyxl_path,
         test_xlsx_mixed_drawing_image_and_textbox, test_xlsx_cell_value_sanitized,
         test_xlsx_data_validation, test_xlsx_drawing_alttext,
         test_xlsx_threaded_comment, test_xlsx_header_footer])
